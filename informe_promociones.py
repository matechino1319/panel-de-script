#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Informe Promociones (Vecinos, Empleados o Jubilados) — Adaptado para Web Panel
# Autor: Adaptado para Gonzapereyraa

import glob
import os
import re
import sys
import unicodedata
from datetime import datetime, timedelta
from calendar import month_name
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from script_runtime import get_input_dir, get_input_file, get_output_dir

# UTF-8 Output
if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except AttributeError:
        import codecs
        sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())

print("=" * 50)
print("GENERADOR DE INFORMES DE PROMOCIONES")
print("=" * 50)

# -------------------------
# BUSCAR ARCHIVO DE ENTRADA
# -------------------------
def encontrar_archivo_entrada():
    archivo_directo = get_input_file()
    if archivo_directo and os.path.exists(archivo_directo):
        return archivo_directo

    input_dir = get_input_dir()
    patrones_excel = ["*.xlsx", "*.xlsm", "*.xls"]
    archivos = []
    for p in patrones_excel:
        archivos.extend(glob.glob(os.path.join(input_dir, p)))
    archivos = sorted(list(dict.fromkeys(archivos)))
    archivos = [a for a in archivos if "Informe_" not in a and not os.path.basename(a).startswith("~$")]

    if archivos:
        return archivos[0]

    archivos_csv = sorted(glob.glob(os.path.join(input_dir, "*.csv")))
    archivos_csv = [a for a in archivos_csv if "Informe_" not in a and not os.path.basename(a).startswith("~$")]

    if archivos_csv:
        return archivos_csv[0]

    return None

archivo_entrada = encontrar_archivo_entrada()
if archivo_entrada is None:
    print("❌ No se encontraron archivos válidos en el directorio:", get_input_dir())
    sys.exit(1)

print(f"✅ Archivo seleccionado automáticamente: {archivo_entrada}")

# -------------------------
# LECTURA DEL ARCHIVO
# -------------------------
ext = os.path.splitext(archivo_entrada)[1].lower()
try:
    if ext in ['.xlsx', '.xls', '.xlsm']:
        df = pd.read_excel(archivo_entrada, sheet_name=0, engine='openpyxl')
    elif ext == '.csv':
        try:
            df = pd.read_csv(archivo_entrada, encoding="utf-8", sep=';', on_bad_lines='skip')
        except Exception:
            df = pd.read_csv(archivo_entrada, encoding="latin-1", sep=';', on_bad_lines='skip')
    else:
        raise ValueError("Formato de archivo no soportado")
    print(f"✅ Archivo leído. Filas: {len(df)}, Columnas: {len(df.columns)}")
except Exception as e:
    print("❌ Error al leer el archivo:", e)
    sys.exit(1)

# -------------------------
# AUTO DETECCIÓN DE GRUPO
# -------------------------
df_sample_str = df.astype(str).to_string().lower()
if "jubilad" in df_sample_str:
    grupo_seleccionado = "JUBILADOS"
    mensajes_filtrar = [
        "jubilado",
        "jubilados",
        "descuento a jubilados",
        "descuento jubilados"
    ]
elif "emplead" in df_sample_str:
    grupo_seleccionado = "EMPLEADOS"
    mensajes_filtrar = [
        "empleado",
        "empleados",
        "descuento a empleados",
        "descuento empleados"
    ]
else:
    grupo_seleccionado = "VECINOS"
    mensajes_filtrar = [
        "10% de descuento a VECINOS",
        "10% de descuento a VECINOS DEBITO",
        "JUEVES 10% de descuento a VECINOS",
        "Miercoles 10% de descuento a VECINOS",
        "10% vecinos",
        "descuento a vecinos",
        "descuento vecinos",
        "vecinos"
    ]

print(f"✅ Grupo detectado automáticamente: {grupo_seleccionado}")

# Configuración de salida
RUTA_DESTINO = get_output_dir(rf"G:\Mi unidad\ARCHIVOS_COMPARTIDOS_LAYUNTA\SISTEMAS\INFORMES {grupo_seleccionado}")

SUCURSALES = [
    "LY01-Ballofet",
    "LY02-Velez",
    "LY04-Alem",
    "LY07-Cuadro Benegas",
    "LY19-Alvear",
    "LY22-CENTRO",
    "LY25-Montoya",
    "LY25-Alberdi",
    "LY34-Bowen",
    "LY37-Libertador",
    "LY42-Atuel Norte"
]

PORCENTAJES_ESPECIALES = {
    "VECINOS":   {"LY44": 0.05},
    "EMPLEADOS": {},
    "JUBILADOS": {"LY44": 0.10},
}

PORCENTAJE_GENERAL = {
    "VECINOS":   0.10,
    "EMPLEADOS": 0.10,
    "JUBILADOS": 0.15,
}

PORCENTAJE_SUCURSAL = {
    "EMPLEADOS": 0.13,
}

def obtener_porcentaje(nombre_tienda, grupo, es_sucursal=False):
    especiales = PORCENTAJES_ESPECIALES.get(grupo, {})
    nombre_upper = str(nombre_tienda).upper()
    for clave, pct in especiales.items():
        if clave.upper() in nombre_upper:
            return pct
    if es_sucursal and grupo in PORCENTAJE_SUCURSAL:
        return PORCENTAJE_SUCURSAL[grupo]
    return PORCENTAJE_GENERAL.get(grupo, 0.10)

# -------------------------
# NORMALIZAR COLUMNAS
# -------------------------
def normalizar_nombre(col):
    coln = unicodedata.normalize('NFKD', str(col))
    coln = coln.encode('ascii', errors='ignore').decode('ascii')
    coln = re.sub(r'[^a-zA-Z0-9]', '_', coln)
    coln = re.sub(r'_+', '_', coln)
    return coln.strip('_')

original_columns = df.columns.tolist()
new_columns = []
for col in original_columns:
    col_str = str(col)
    if col_str == '$Beneficio':
        col_normalized = 'Beneficio_monetario'
    elif col_str == '%Beneficio':
        col_normalized = 'Beneficio_porcentaje'
    elif col_str == 'Beneficio':
        col_normalized = 'Beneficio_general'
    else:
        col_normalized = normalizar_nombre(col_str)
    base_name = col_normalized
    counter = 1
    while col_normalized in new_columns:
        col_normalized = f"{base_name}_{counter}"
        counter += 1
    new_columns.append(col_normalized)
df.columns = new_columns

def encontrar_columna(patrones):
    for col in df.columns:
        col_lower = col.lower()
        for patron in patrones:
            if patron in col_lower:
                return col
    return None

col_beneficio = None
prioritarias = ['beneficio_monetario', '$beneficio', 'dbeneficio', 'beneficio_$', 'monto_beneficio']
for col in df.columns:
    cl = col.lower()
    for p in prioritarias:
        if p in cl:
            col_beneficio = col
            break
    if col_beneficio:
        break
if col_beneficio is None:
    for col in df.columns:
        if 'beneficio' in col.lower() and 'porcentaje' not in col.lower():
            col_beneficio = col
            break
if col_beneficio is None:
    col_beneficio = encontrar_columna(['monto', 'valor', 'importe', 'descuento', 'discount'])

col_mensaje = encontrar_columna(['mensaje', 'promocion', 'promoción', 'promo'])
col_nro_trx = encontrar_columna(['nro', 'trx', 'transaccion', 'numero', 'nro_trx'])
col_fecha = encontrar_columna(['fecha', 'inicio', 'date', 'fecha_inicio_trx'])
col_tienda = encontrar_columna(['tienda', 'sucursal', 'store', 'local'])
col_pct_beneficio = encontrar_columna(['beneficio_porcentaje', '%beneficio', 'porcentaje'])

if col_nro_trx is None:
    for col in df.columns:
        if 'nro' in col.lower() and 'fecha' not in col.lower():
            col_nro_trx = col
            break
if col_fecha is None:
    for col in df.columns:
        if 'fecha' in col.lower():
            col_fecha = col
            break

columnas_requeridas = {
    'Mensaje (Promoción)': col_mensaje,
    'NroTrx': col_nro_trx,
    'Fecha': col_fecha,
    'Beneficio': col_beneficio,
    'Tienda': col_tienda
}

faltantes = [k for k, v in columnas_requeridas.items() if v is None]
if faltantes:
    print("❌ No se encontraron columnas requeridas:", faltantes)
    print("Columnas disponibles:", df.columns.tolist())
    sys.exit(1)

# -------------------------
# FILTRAR FILAS
# -------------------------
patron_mensajes = "|".join([re.escape(m).lower() for m in mensajes_filtrar])
df_filtrado = df[df[col_mensaje].astype(str).str.lower().str.contains(patron_mensajes, na=False)].copy()

if df_filtrado.empty:
    print(f"⚠ No se encontraron filas que coincidan con las promociones buscadas para {grupo_seleccionado}.")
    sys.exit(0)

print(f"✅ Filtrado promociones para {grupo_seleccionado}: {len(df_filtrado)} filas")

def limpiar_valor_monetario(valor):
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip()
    if s == '' or s.lower() in ['nan', 'none']:
        return 0.0
    s = re.sub(r'[^\d,-.]', '', s)
    if ',' in s and '.' in s:
        if s.find(',') > s.find('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    else:
        s = s.replace(',', '.')
    if s.count('.') > 1:
        parts = s.split('.')
        s = ''.join(parts[:-1]) + '.' + parts[-1]
    try:
        return abs(float(s))
    except Exception:
        return 0.0

df_filtrado['Beneficio_limpio'] = df_filtrado[col_beneficio].apply(limpiar_valor_monetario)

if col_pct_beneficio:
    def resolver_porcentaje_fila(row):
        val = limpiar_valor_monetario(row[col_pct_beneficio])
        if val and val > 0:
            return val / 100 if val > 1 else val
        return obtener_porcentaje(row[col_tienda], grupo_seleccionado, es_sucursal=(row[col_tienda] in SUCURSALES))
    df_filtrado['Porcentaje_descuento'] = df_filtrado.apply(resolver_porcentaje_fila, axis=1)
else:
    df_filtrado['Porcentaje_descuento'] = df_filtrado.apply(
        lambda row: obtener_porcentaje(row[col_tienda], grupo_seleccionado, es_sucursal=(row[col_tienda] in SUCURSALES)),
        axis=1
    )

df_filtrado['Venta_Neta_fila'] = df_filtrado['Beneficio_limpio'] / df_filtrado['Porcentaje_descuento']

df_actual = df_filtrado.copy()
df_sucursales = df_actual[df_actual[col_tienda].isin(SUCURSALES)].copy()
df_franquicias = df_actual[~df_actual[col_tienda].isin(SUCURSALES)].copy()

def crear_hoja_excel(ws, df_data, titulo_hoja, tipo, periodo_text="PERIODO COMPLETO"):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
    if tipo == "sucursal":
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        total_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    else:
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    title_font = Font(bold=True, size=16, color="FFFFFF")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    total_font = Font(bold=True, size=11, color="000000")
    normal_font = Font(size=11)

    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    resumen = df_data.groupby(col_tienda).agg({
        col_nro_trx: pd.Series.nunique,
        'Beneficio_limpio': "sum",
        'Venta_Neta_fila': "sum",
        'Porcentaje_descuento': "first"
    }).reset_index()

    resumen.rename(columns={
        col_nro_trx: "Transacciones totales",
        'Beneficio_limpio': "Total Descuento",
        'Venta_Neta_fila': "Venta Neta Total",
        'Porcentaje_descuento': "% Descuento"
    }, inplace=True)

    ws.merge_cells('A1:E1')
    ws['A1'] = f"{titulo_hoja} - {periodo_text}"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws['A1'].fill = header_fill

    ws.merge_cells('A2:E2')
    tipo_texto = "Sucursales" if tipo == "sucursal" else "Franquicias"
    ws['A2'] = f"{tipo_texto} - Total Tiendas: {len(resumen)} | Transacciones: {resumen['Transacciones totales'].sum():,} | Período: {periodo_text}"
    ws['A2'].font = Font(bold=True, size=10, color="000000")
    ws['A2'].alignment = center_align
    ws['A2'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.append([])
    columnas_excel = ["Tienda", "Transacciones totales", "Total Descuento", "Venta Neta Total", "% Descuento"]
    ws.append(columnas_excel)

    for col_idx in range(1, 6):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    fila_inicio = 5
    for idx, row in resumen.iterrows():
        ws.cell(row=fila_inicio, column=1, value=row[col_tienda]).alignment = left_align
        ws.cell(row=fila_inicio, column=2, value=int(row["Transacciones totales"])).alignment = center_align
        ws.cell(row=fila_inicio, column=3, value=row["Total Descuento"])
        ws.cell(row=fila_inicio, column=4, value=row["Venta Neta Total"])
        pct_cell = ws.cell(row=fila_inicio, column=5, value=row["% Descuento"])
        pct_cell.number_format = '0.00%'
        pct_cell.alignment = center_align
        fila_inicio += 1

    for row in ws.iter_rows(min_row=5, max_row=fila_inicio-1, min_col=1, max_col=5):
        for cell in row:
            cell.border = thin_border
            cell.font = normal_font
            if cell.column in [3, 4]:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = right_align
            elif cell.column == 2:
                cell.alignment = center_align
            elif cell.column == 5:
                cell.number_format = '0.00%'
                cell.alignment = center_align

    ws.cell(row=fila_inicio, column=1, value="TOTAL GENERAL").font = total_font
    ws.cell(row=fila_inicio, column=2, value=int(resumen["Transacciones totales"].sum())).font = total_font
    ws.cell(row=fila_inicio, column=3, value=resumen["Total Descuento"].sum()).font = total_font
    ws.cell(row=fila_inicio, column=4, value=resumen["Venta Neta Total"].sum()).font = total_font
    ws.cell(row=fila_inicio, column=5, value="")

    for col_idx in range(1, 6):
        cell = ws.cell(row=fila_inicio, column=col_idx)
        cell.fill = total_fill
        cell.border = thick_border
        cell.alignment = center_align
        if col_idx in [3, 4]:
            cell.number_format = '"$"#,##0.00'

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 14

    return resumen

# -------------------------
# GENERAR EXCEL FINAL
# -------------------------
wb = Workbook()
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

periodo_text = "PERIODO COMPLETO (SIN FILTRADO POR FECHAS)"

ws_suc = wb.create_sheet("Sucursales")
res_suc = crear_hoja_excel(ws_suc, df_sucursales, f"INFORME {grupo_seleccionado} - SUCURSALES", "sucursal", periodo_text)

ws_fran = wb.create_sheet("Franquicias")
res_fran = crear_hoja_excel(ws_fran, df_franquicias, f"INFORME {grupo_seleccionado} - FRANQUICIAS", "franquicia", periodo_text)

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
nombre_archivo = f"Informe_{grupo_seleccionado}_{timestamp}.xlsx"
ruta_completa = os.path.join(RUTA_DESTINO, nombre_archivo)

os.makedirs(RUTA_DESTINO, exist_ok=True)
wb.save(ruta_completa)
print(f"✅ Informe promociones generado exitosamente: {ruta_completa}")
