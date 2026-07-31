import glob
import os
import sys

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from script_runtime import get_input_dir, get_input_file, get_output_dir


if sys.stdout.encoding != "utf-8":
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except AttributeError:
        import codecs

        sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())


folder_path = get_input_dir(os.getcwd())


def encontrar_archivo_entrada():
    archivo_directo = get_input_file()
    if archivo_directo and os.path.exists(archivo_directo):
        return archivo_directo

    patrones = ["*.xlsx", "*.xls", "*.xlsm", "*.csv", "*.XLSX", "*.XLS", "*.XLSM", "*.CSV"]
    archivos_encontrados = []
    for patron in patrones:
        archivos = glob.glob(os.path.join(folder_path, patron))
        archivos = [f for f in archivos if not os.path.basename(f).startswith("~$")]
        archivos_encontrados.extend(archivos)

    if not archivos_encontrados:
        return None

    return max(archivos_encontrados, key=os.path.getctime)


def leer_archivo(path_archivo):
    extension = os.path.splitext(path_archivo)[1].lower()
    if extension in {".xlsx", ".xls", ".xlsm"}:
        return pd.read_excel(path_archivo)

    if extension == ".csv":
        errores = []
        for encoding in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
            try:
                return pd.read_csv(path_archivo, sep=";", encoding=encoding, on_bad_lines="skip")
            except Exception as exc:
                errores.append(f"{encoding}: {exc}")
        raise ValueError("No se pudo leer el CSV. " + " | ".join(errores))

    raise ValueError(f"Formato no soportado: {extension}")


input_file = encontrar_archivo_entrada()

if not input_file:
    print("No se encontro ningun archivo compatible en la carpeta.")
    sys.exit(1)

print(f"Archivo detectado: {os.path.basename(input_file)}")

df = leer_archivo(input_file)
print(f"Total de filas cargadas: {len(df)}")
print(f"Columnas: {df.columns.tolist()}")

promo_col = None
for col in df.columns:
    if "promoci" in str(col).lower():
        promo_col = col
        break

if promo_col is None:
    print("Columnas disponibles:", df.columns.tolist())
    raise ValueError("No se encontro la columna de Promocion")

jubilados_mask = df[promo_col].astype(str).str.contains("JUBILADO", case=False, na=False)
print(f"Filas con 'JUBILADO': {jubilados_mask.sum()}")
if jubilados_mask.sum() > 0:
    print("Ejemplos de promociones con JUBILADO:")
    print(df[jubilados_mask][promo_col].dropna().unique()[:10])

df_filtrado = df[jubilados_mask].copy()

print(f"\nFilas despues del filtro: {len(df_filtrado)}")
if len(df_filtrado) == 0:
    print(f"\nNo se encontraron datos con 'JUBILADO' en la columna '{promo_col}'")
    print("\nTodas las promociones unicas en el archivo:")
    print(df[promo_col].dropna().unique())

sucursales = [
    "LY01-Ballofet",
    "LY02-Velez",
    "LY04-Alem",
    "LY07-Cuadro Benegas",
    "LY22-CENTRO",
    "LY37-Libertador",
    "LY34-Bowen",
    "LY19-Alvear",
    "LY42-Atuel Norte",
    "LY41-DEPOSITODANI",
    "LY24-Deposito Logistica y Distribucion",
]


def detectar_columna(df_local, opciones):
    normalizadas = {str(col).strip().lower(): col for col in df_local.columns}
    for opcion in opciones:
        if opcion in normalizadas:
            return normalizadas[opcion]
    for col in df_local.columns:
        nombre = str(col).strip().lower()
        if any(opcion in nombre for opcion in opciones):
            return col
    return None


tienda_col = detectar_columna(df_filtrado, ["tienda", "store", "sucursal", "local"])
trx_col = detectar_columna(df_filtrado, ["nro trx", "nro_trx", "trx", "transaccion"])
beneficio_col = None
for candidata in ["$beneficio", "beneficio", "importe", "monto", "descuento"]:
    beneficio_col = detectar_columna(df_filtrado, [candidata])
    if beneficio_col:
        break

faltantes = []
if tienda_col is None:
    faltantes.append("Tienda")
if trx_col is None:
    faltantes.append("Nro Trx")
if beneficio_col is None:
    faltantes.append("Beneficio")

if faltantes:
    raise ValueError("No se encontraron columnas requeridas: " + ", ".join(faltantes))


def limpiar_valor_monetario(valor):
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)

    texto = str(valor).strip()
    if not texto:
        return 0.0

    texto = "".join(ch for ch in texto if ch.isdigit() or ch in ",.-")
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            texto = texto.replace(".", "").replace(",", ".")
        else:
            texto = texto.replace(",", "")
    else:
        texto = texto.replace(",", ".")

    if texto.count(".") > 1:
        partes = texto.split(".")
        texto = "".join(partes[:-1]) + "." + partes[-1]

    try:
        return float(texto)
    except ValueError:
        return 0.0


df_filtrado["Tipo"] = df_filtrado[tienda_col].apply(
    lambda tienda: "Sucursal" if tienda in sucursales else "Franquicia"
)
df_filtrado["BeneficioLimpio"] = df_filtrado[beneficio_col].apply(limpiar_valor_monetario)

tabla = (
    df_filtrado.groupby(["Tipo", tienda_col])
    .agg(Transacciones=(trx_col, "nunique"), Descuento=("BeneficioLimpio", "sum"))
    .reset_index()
    .rename(columns={tienda_col: "Tienda"})
)

tabla_sucursal = tabla[tabla["Tipo"] == "Sucursal"].copy()
tabla_franquicia = tabla[tabla["Tipo"] == "Franquicia"].copy()

total_sucursal = pd.DataFrame(
    {
        "Tipo": [""],
        "Tienda": ["TOTAL"],
        "Transacciones": [tabla_sucursal["Transacciones"].sum()],
        "Descuento": [tabla_sucursal["Descuento"].sum()],
    }
)
total_franquicia = pd.DataFrame(
    {
        "Tipo": [""],
        "Tienda": ["TOTAL"],
        "Transacciones": [tabla_franquicia["Transacciones"].sum()],
        "Descuento": [tabla_franquicia["Descuento"].sum()],
    }
)

final_sucursal = pd.concat([tabla_sucursal, total_sucursal], ignore_index=True)
final_franquicia = pd.concat([tabla_franquicia, total_franquicia], ignore_index=True)

for df_final in [final_sucursal, final_franquicia]:
    df_final["Descuento"] = df_final["Descuento"].map(
        lambda x: f"${x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    )

espacio = pd.DataFrame({"Tipo": [""], "Tienda": [""], "Transacciones": [""], "Descuento": [""]})
sucursal_header = pd.DataFrame(
    {"Tipo": ["Sucursal"], "Tienda": [""], "Transacciones": [""], "Descuento": [""]}
)
franquicia_header = pd.DataFrame(
    {"Tipo": ["Franquicia"], "Tienda": [""], "Transacciones": [""], "Descuento": [""]}
)

full_report = pd.concat(
    [sucursal_header, final_sucursal, espacio, franquicia_header, final_franquicia],
    ignore_index=True,
)

output_file = os.path.join(get_output_dir(folder_path), "REPORTE_PROMOCIONES_JUBILADOS.xlsx")
full_report.to_excel(output_file, sheet_name="REPORTE", index=False)

wb = load_workbook(output_file)
ws = wb["REPORTE"]

header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
row_fill_odd = PatternFill(start_color="E3F1FA", end_color="E3F1FA", fill_type="solid")
row_fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
total_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
total_font = Font(bold=True, color="333333", size=12)
section_fill = PatternFill(start_color="005D98", end_color="005D98", fill_type="solid")
section_font = Font(bold=True, color="FFFFFF", size=14)
center_align = Alignment(horizontal="center", vertical="center")
title_font = Font(bold=True, color="005D98", size=16)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

ws.insert_rows(1)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
ws["A1"] = "REPORTE PROMOCIONES JUBILADOS"
ws["A1"].font = title_font
ws["A1"].alignment = center_align

for cell in ws[2]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

for idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
    tipo = row[0].value
    tienda = row[1].value
    if tipo in ["Sucursal", "Franquicia"]:
        for cell in row:
            cell.font = section_font
            cell.fill = section_fill
            cell.alignment = center_align
            cell.border = thin_border
    elif tienda == "TOTAL":
        for cell in row:
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = center_align
            cell.border = thin_border
    else:
        fill = row_fill_odd if idx % 2 == 1 else row_fill_even
        for cell in row:
            cell.font = Font(size=11)
            cell.fill = fill
            cell.alignment = center_align
            cell.border = thin_border

for column_cells in ws.columns:
    length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 5

wb.save(output_file)
print("Reporte generado exitosamente en:", output_file)
