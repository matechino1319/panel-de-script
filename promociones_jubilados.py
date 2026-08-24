import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from script_runtime import get_input_file, get_output_dir

input_file = get_input_file()
if not input_file:
    raise FileNotFoundError("No se recibió ningún archivo de entrada.")

if input_file.lower().endswith('.csv'):
    try:
        df = pd.read_csv(input_file, encoding="utf-8", sep=';', on_bad_lines='skip')
    except Exception:
        df = pd.read_csv(input_file, encoding="latin-1", sep=';', on_bad_lines='skip')
else:
    df = pd.read_excel(input_file)

def find_col(df, keywords):
    for col in df.columns:
        if any(k in col.lower() for k in keywords):
            return col
    raise KeyError(f"No se encontró columna para {keywords}. Columnas: {list(df.columns)}")

col_promo = find_col(df, ['promo', 'mensaje'])
col_tienda = find_col(df, ['tienda', 'sucursal'])
col_trx = find_col(df, ['trx', 'transaccion', 'nro'])
col_beneficio = find_col(df, ['beneficio', 'descuento', 'monto'])

# Limpiar beneficio por si viene como texto con $ o comas
def limpiar_valor_monetario(valor):
    if pd.isna(valor): return 0.0
    if isinstance(valor, (int, float)): return float(valor)
    import re
    s = str(valor).strip()
    s = re.sub(r'[^\d,-.]', '', s)
    if ',' in s and '.' in s:
        if s.find(',') > s.find('.'): s = s.replace('.', '').replace(',', '.')
        else: s = s.replace(',', '')
    else: s = s.replace(',', '.')
    if s.count('.') > 1:
        parts = s.split('.')
        s = ''.join(parts[:-1]) + '.' + parts[-1]
    try: return abs(float(s))
    except: return 0.0

df[col_beneficio] = df[col_beneficio].apply(limpiar_valor_monetario)

# Filtrar solo promociones de jubilados lunes y martes
df_filtrado = df[df[col_promo].astype(str).str.contains('JUBILADOS', case=False, na=False) & 
                 df[col_promo].astype(str).str.contains('LUNES|MARTES', case=False, na=False)].copy()

# Lista de tiendas sucursal
sucursales = [
    "LY01-Ballofet", "LY02-Velez", "LY04-Alem", "LY07-Cuadro Benegas", "LY22-CENTRO",
    "LY37-Libertador", "LY25-Montoya", "LY19-Alvear", "LY42-Atuel Norte", "LY41-DEPOSITODANI",
    "LY24-Deposito Logistica y Distribucion", "LY25-Alberdi"
]

# Crear columna Tipo
df_filtrado['Tipo'] = df_filtrado[col_tienda].apply(lambda tienda: 'Sucursal' if tienda in sucursales else 'Franquicia')

# Agrupar por Tipo y Tienda, contando transacciones únicas y sumando descuento
tabla = (
    df_filtrado.groupby(['Tipo', col_tienda])
    .agg(Transacciones=(col_trx, 'nunique'), Descuento=(col_beneficio, 'sum'))
    .reset_index()
).rename(columns={col_tienda: 'Tienda'})

# Separar Sucursales y Franquicias
tabla_sucursal = tabla[tabla['Tipo'] == 'Sucursal'].copy()
tabla_franquicia = tabla[tabla['Tipo'] == 'Franquicia'].copy()

# Totales
total_sucursal = pd.DataFrame({
    'Tipo': [''],
    'Tienda': ['TOTAL'],
    'Transacciones': [tabla_sucursal['Transacciones'].sum()],
    'Descuento': [tabla_sucursal['Descuento'].sum()]
})
total_franquicia = pd.DataFrame({
    'Tipo': [''],
    'Tienda': ['TOTAL'],
    'Transacciones': [tabla_franquicia['Transacciones'].sum()],
    'Descuento': [tabla_franquicia['Descuento'].sum()]
})

# Juntar tablas
final_sucursal = pd.concat([tabla_sucursal, total_sucursal], ignore_index=True)
final_franquicia = pd.concat([tabla_franquicia, total_franquicia], ignore_index=True)

# Formato moneda
for df_final in [final_sucursal, final_franquicia]:
    df_final['Descuento'] = df_final['Descuento'].map(lambda x: f"${x:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

# Unir las dos tablas en una sola hoja, con encabezados de sección
espacio = pd.DataFrame({'Tipo': [''], 'Tienda': [''], 'Transacciones': [''], 'Descuento': ['']}, index=[0])
sucursal_header = pd.DataFrame({'Tipo': ['Sucursal'], 'Tienda': [''], 'Transacciones': [''], 'Descuento': ['']}, index=[0])
franquicia_header = pd.DataFrame({'Tipo': ['Franquicia'], 'Tienda': [''], 'Transacciones': [''], 'Descuento': ['']}, index=[0])

full_report = pd.concat([
    sucursal_header,
    final_sucursal,
    espacio,
    franquicia_header,
    final_franquicia
], ignore_index=True)

# Escribir el reporte en Excel con formato
output_file = os.path.join(get_output_dir(), 'REPORTE_PROMOCIONES_JUBILADOS.xlsx')
full_report.to_excel(output_file, sheet_name='REPORTE', index=False)

# Formatear con openpyxl
wb = load_workbook(output_file)
ws = wb['REPORTE']

# Estilos
header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")  # Azul oscuro
header_font = Font(bold=True, color="FFFFFF", size=12)
row_fill_odd = PatternFill(start_color="E3F1FA", end_color="E3F1FA", fill_type="solid")  # Azul claro
row_fill_even = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid") # Blanco
total_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")    # Dorado
total_font = Font(bold=True, color="333333", size=12)
section_fill = PatternFill(start_color="005D98", end_color="005D98", fill_type="solid")  # Azul intermedio
section_font = Font(bold=True, color="FFFFFF", size=14)
center_align = Alignment(horizontal="center", vertical="center")
title_font = Font(bold=True, color="005D98", size=16)
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                     top=Side(style='thin'), bottom=Side(style='thin'))

# Título principal
ws.insert_rows(1)
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
ws['A1'] = "REPORTE PROMOCIONES JUBILADOS LUNES Y MARTES"
ws['A1'].font = title_font
ws['A1'].alignment = center_align

# Formato encabezados de tabla
for cell in ws[2]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

# Formato filas alternas y totales
for idx, row in enumerate(ws.iter_rows(min_row=3, max_row=ws.max_row), start=3):
    tipo = row[0].value
    tienda = row[1].value
    # Encabezados de sección
    if tipo in ["Sucursal", "Franquicia"]:
        for cell in row:
            cell.font = section_font
            cell.fill = section_fill
            cell.alignment = center_align
            cell.border = thin_border
    # Totales
    elif tienda == "TOTAL":
        for cell in row:
            cell.font = total_font
            cell.fill = total_fill
            cell.alignment = center_align
            cell.border = thin_border
    # Filas normales alternas
    else:
        fill = row_fill_odd if idx % 2 == 1 else row_fill_even
        for cell in row:
            cell.font = Font(size=11)
            cell.fill = fill
            cell.alignment = center_align
            cell.border = thin_border

# Autoajustar ancho de columnas
for column_cells in ws.columns:
    length = max(len(str(cell.value)) if cell.value else 0 for cell in column_cells)
    ws.column_dimensions[get_column_letter(column_cells[0].column)].width = length + 5

wb.save(output_file)
print("Reporte generado exitosamente en:", output_file)