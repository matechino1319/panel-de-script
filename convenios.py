import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from calendar import month_name
from datetime import datetime, timedelta
import re
import unicodedata
import os
import glob
from script_runtime import get_input_dir, get_output_dir, get_input_file

# =========================
# CONFIGURACIÓN - BÚSQUEDA AUTOMÁTICA DE ARCHIVO CSV
# =========================
def encontrar_archivo_csv():
    """
    Busca automáticamente archivos CSV en el directorio actual.
    Retorna el nombre del primer archivo CSV encontrado.
    """
    archivo_directo = get_input_file()
    if archivo_directo and archivo_directo.lower().endswith('.csv'):
        print(f"Archivo subido detectado: {os.path.basename(archivo_directo)}")
        return archivo_directo

    # Buscar todos los archivos CSV en el directorio configurado
    archivos_csv = glob.glob(os.path.join(get_input_dir(), "*.csv"))
    
    if not archivos_csv:
        print("❌ No se encontraron archivos CSV en el directorio actual")
        print("📁 Directorio actual:", os.getcwd())
        return None
    
    if len(archivos_csv) == 1:
        print(f"✅ Se encontró 1 archivo CSV: {archivos_csv[0]}")
        return archivos_csv[0]
    else:
        print("📂 Se encontraron múltiples archivos CSV:")
        for i, archivo in enumerate(archivos_csv, 1):
            print(f"   {i}. {archivo}")
        
        while True:
            try:
                seleccion = input(f"\nSeleccione el archivo a procesar (1-{len(archivos_csv)}): ")
                seleccion = int(seleccion) - 1
                if 0 <= seleccion < len(archivos_csv):
                    return archivos_csv[seleccion]
                else:
                    print(f"Por favor, ingrese un número entre 1 y {len(archivos_csv)}")
            except ValueError:
                print("Por favor, ingrese un número válido")

# Buscar automáticamente el archivo CSV
archivo_entrada = encontrar_archivo_csv()
if archivo_entrada is None:
    exit()

# Ruta de destino en Google Drive compartido
RUTA_DESTINO = get_output_dir(r"G:\Mi unidad\ARCHIVOS_COMPARTIDOS_LAYUNTA\SISTEMAS\INFORMES VECINOS")

# Lista de sucursales (las demás serán consideradas franquicias)
SUCURSALES = [
    "LY01-Ballofet",
    "LY02-Velez", 
    "LY04-Alem",
    "LY07-Cuadro Benegas",
    "LY19-Alvear",
    "LY22-CENTRO",
    "LY34-Bowen",
    "LY37-Libertador",
    "LY42-Atuel Norte"
]

# =========================
# FUNCIÓN PARA DETERMINAR SEMANAS BASADAS EN MIÉRCOLES Y JUEVES
# =========================
def obtener_semana_miercoles_jueves(fecha):
    """
    Determina la semana del mes basado en miércoles y jueves consecutivos.
    Cada semana contiene un miércoles y el jueves siguiente.
    """
    if pd.isna(fecha):
        return 0
        
    # Encontrar el primer miércoles del mes
    primer_dia_mes = datetime(fecha.year, fecha.month, 1)
    dias_hasta_primer_miercoles = (2 - primer_dia_mes.weekday()) % 7
    primer_miercoles = primer_dia_mes + timedelta(days=dias_hasta_primer_miercoles)
    
    # Calcular la semana basada en la distancia desde el primer miércoles
    dias_diferencia = (fecha - primer_miercoles).days
    semana = (dias_diferencia // 7) + 1
    
    # Ajustar para fechas antes del primer miércoles
    if dias_diferencia < 0:
        return 0
    
    # Asegurar que no exceda la semana 5
    return min(semana, 5)

def obtener_rango_semana_miercoles_jueves(mes, anio, numero_semana):
    """
    Obtiene el rango de fechas para una semana específica basada en miércoles y jueves.
    """
    if numero_semana == 0:
        # Para días antes de la primera semana
        primer_dia_mes = datetime(anio, mes, 1)
        dias_hasta_primer_miercoles = (2 - primer_dia_mes.weekday()) % 7
        primer_miercoles = primer_dia_mes + timedelta(days=dias_hasta_primer_miercoles)
        fecha_inicio = primer_dia_mes
        fecha_fin = primer_miercoles - timedelta(days=1)
        return fecha_inicio, fecha_fin
    
    # Encontrar el primer miércoles del mes
    primer_dia_mes = datetime(anio, mes, 1)
    dias_hasta_primer_miercoles = (2 - primer_dia_mes.weekday()) % 7
    primer_miercoles = primer_dia_mes + timedelta(days=dias_hasta_primer_miercoles)
    
    # Calcular el miércoles de la semana solicitada
    miercoles_semana = primer_miercoles + timedelta(days=(numero_semana - 1) * 7)
    jueves_semana = miercoles_semana + timedelta(days=1)
    
    return miercoles_semana, jueves_semana

# =========================
# SOLICITAR SEMANA A ANALIZAR
# =========================
# =========================
# CONFIGURACIÓN - ANALIZAR TODAS LAS FECHAS
# =========================
# No filtrar por semana, analizar todos los datos
semana_analizar = 0  # O cualquier valor por defecto
mes_analizar = datetime.today().month
anio_analizar = datetime.today().year

print(f"\n📊 Analizando TODOS los datos de convenios (sin filtro por semana)")

# =========================
# LECTURA DEL CSV CON SEPARADOR CORRECTO
# =========================
try:
    # Leer el CSV con separador punto y coma
    df = pd.read_csv(archivo_entrada, encoding="utf-8", sep=';', on_bad_lines='skip')
    print(f"✅ CSV leído exitosamente. Filas: {len(df)}, Columnas: {len(df.columns)}")
except Exception as e:
    print(f"❌ Error al leer el CSV: {e}")
    # Intentar con otros encodings
    try:
        df = pd.read_csv(archivo_entrada, encoding="latin-1", sep=';', on_bad_lines='skip')
        print("✅ CSV leído con encoding latin-1")
    except Exception as e2:
        print(f"❌ No se pudo leer el CSV: {e2}")
        exit()

# Mostrar información del DataFrame
print("Columnas originales:", df.columns.tolist())

# =========================
# DIAGNÓSTICO DE COLUMNAS MONETARIAS
# =========================
print("\n=== DIAGNÓSTICO DE COLUMNAS ===")
print("DataFrame shape:", df.shape)
print("Tipos de datos:")
print(df.dtypes)



# Buscar específicamente columnas monetarias
monetary_columns = []
for col in df.columns:
    if any(term in col.lower() for term in ['beneficio', 'monto', 'valor', 'importe', 'descuento', '$']):
        print(f"\nColumna monetaria candidata: {col}")
        print(f"Tipo: {df[col].dtype}")
        print("Muestra de valores:", df[col].unique()[:10])
        monetary_columns.append(col)

print(f"\nColumnas monetarias identificadas: {monetary_columns}")

# =========================
# NORMALIZAR NOMBRES DE COLUMNAS (SIMPLIFICADA)
# =========================
# Crear nuevos nombres de columnas únicos manteniendo referencia to las originales
original_columns = df.columns.tolist()
new_columns = []

for i, col in enumerate(original_columns):
    # Crear nombre normalizado pero mantener información del original
    if col == '$Beneficio':
        col_normalized = 'Beneficio_monetario'
    elif col == '%Beneficio':
        col_normalized = 'Beneficio_porcentaje'
    elif col == 'Beneficio':
        col_normalized = 'Beneficio_general'
    else:
        # Normalizar otros nombres
        col_normalized = unicodedata.normalize('NFKD', str(col))
        col_normalized = col_normalized.encode('ascii', errors='ignore').decode('ascii')
        col_normalized = re.sub(r'[^a-zA-Z0-9]', '_', col_normalized)
        col_normalized = re.sub(r'_+', '_', col_normalized)
        col_normalized = col_normalized.strip('_')
    
    # Hacer el nombre único si ya existe
    base_name = col_normalized
    counter = 1
    while col_normalized in new_columns:
        col_normalized = f"{base_name}_{counter}"
        counter += 1
    
    new_columns.append(col_normalized)

df.columns = new_columns

print("Columnas normalizadas:", df.columns.tolist())

# =========================
# IDENTIFICAR COLUMNAS CORRECTAS - CORREGIDO
# =========================
# Buscar columnas por patrones
def encontrar_columna(patrones, excluir=None):
    for col in df.columns:
        col_lower = col.lower()
        if excluir and any(exc in col_lower for exc in excluir):
            continue
        for patron in patrones:
            if patron in col_lower:
                return col
    return None

# MEJORADA: Búsqueda prioritaria de columna monetaria
col_beneficio = None
columnas_prioritarias = ['beneficio_monetario', '$beneficio', 'dbeneficio', 'beneficio_$', 'monto_beneficio']

for col in df.columns:
    col_lower = col.lower()
    for prioritaria in columnas_prioritarias:
        if prioritaria in col_lower:
            col_beneficio = col
            break
    if col_beneficio:
        break

# Si no encontramos, buscar cualquier columna con "beneficio" excepto porcentaje
if col_beneficio is None:
    for col in df.columns:
        if 'beneficio' in col.lower() and 'porcentaje' not in col.lower():
            col_beneficio = col
            break

# Si aún no encontramos, buscar columnas monetarias genéricas
if col_beneficio is None:
    col_beneficio = encontrar_columna(['monto', 'valor', 'importe', 'descuento', 'discount'])

# Identificar otras columnas importantes con mayor precisión
col_mensaje = encontrar_columna(['mensaje', 'promocion'])
# Para Nro Trx, excluir 'fecha' y 'tipo' para evitar confusión con 'Fecha Inicio Trx' o 'Tipo Trx'
col_nro_trx = encontrar_columna(['nro_trx', 'nro', 'trx', 'transaccion', 'numero'], excluir=['fecha', 'tipo'])
col_fecha = encontrar_columna(['fecha', 'inicio', 'date', 'fecha_inicio_trx'])
col_tienda = encontrar_columna(['tienda', 'sucursal', 'store', 'local'])

# Verificar que todas las columnas necesarias se encontraron
columnas_requeridas = {
    'Mensaje': col_mensaje,
    'NroTrx': col_nro_trx,
    'Fecha': col_fecha,
    'Beneficio ($)': col_beneficio,
    'Tienda': col_tienda
}

print(f"🔍 Columna mensaje identificada: {col_mensaje}")
print(f"🔍 Columna nro trx identificada: {col_nro_trx}")
print(f"🔍 Columna fecha identificada: {col_fecha}")
print(f"🔍 Columna beneficio ($) identificada: {col_beneficio}")
print(f"🔍 Columna tienda identificada: {col_tienda}")

# Verificación adicional de la columna de beneficio
if col_beneficio:
    print(f"\n🔍 Verificación de columna de beneficio ({col_beneficio}):")
    # print(f"Muestra de datos: {df[col_beneficio].head(10).tolist()}") # Comentado para evitar mucho output
    print(f"Tipo de datos: {df[col_beneficio].dtype}")
    print(f"Valores no nulos: {df[col_beneficio].notnull().sum()}")

columnas_faltantes = [k for k, v in columnas_requeridas.items() if v is None]
if columnas_faltantes:
    print(f"❌ No se encontraron las siguientes columnas: {columnas_faltantes}")
    print("Columnas disponibles:", df.columns.tolist())
    exit()

# =========================
# FILTRADO PROMOCIONES - DESCUENTOS A VECINOS
# =========================

# Filtrar promociones
# Filtrar solo mensajes que contengan "convenio"
df_filtrado = df[df[col_mensaje].str.contains('convenio', case=False, na=False)].copy()

if df_filtrado.empty:
    print("⚠ No hay datos que coincidan con los mensajes de promoción")
    exit()

print(f"✅ Datos filtrados: {len(df_filtrado)} filas encontradas")

# =========================
# PROCESAMIENTO DE DATOS
# =========================
try:
    # MEJORADA: Función para limpiar y convertir valores monetarios
    def limpiar_valor_monetario(valor):
        if pd.isna(valor):
            return 0.0
        
        # Si ya es numérico, devolver directamente
        if isinstance(valor, (int, float)):
            return float(valor)
        
        valor_str = str(valor).strip()
        
        if valor_str == '' or valor_str.lower() == 'nan' or valor_str.lower() == 'none':
            return 0.0
        
        # Eliminar símbolos de moneda y espacios
        valor_str = re.sub(r'[^\d,-.]', '', valor_str)
        
        # Reemplazar coma decimal por punto
        if ',' in valor_str and '.' in valor_str:
            # Si hay ambos, determinar cuál es el separador decimal
            if valor_str.find(',') > valor_str.find('.'):
                valor_str = valor_str.replace('.', '').replace(',', '.')
            else:
                valor_str = valor_str.replace(',', '')
        else:
            valor_str = valor_str.replace(',', '.')
        
        # Eliminar múltiples puntos decimales
        if valor_str.count('.') > 1:
            partes = valor_str.split('.')
            parte_entera = ''.join(partes[:-1])
            parte_decimal = partes[-1]
            valor_str = f"{parte_entera}.{parte_decimal}"
        
        try:
            return abs(float(valor_str))  # Usar valor absoluto para descuentos
        except ValueError:
            print(f"⚠ No se pudo convertir: '{valor}' a numérico")
            return 0.0
    
    # Aplicar la limpieza a la columna de beneficio
    df_filtrado['Beneficio_limpio'] = df_filtrado[col_beneficio].apply(limpiar_valor_monetario)
    
    # Procesar fechas
    df_filtrado[col_fecha] = pd.to_datetime(df_filtrado[col_fecha], errors="coerce", dayfirst=True)
    
    # Eliminar filas con fechas inválidas
    df_filtrado = df_filtrado.dropna(subset=[col_fecha])
        
    print("✅ Datos procesados correctamente")

except Exception as e:
    print(f"❌ Error al procesar datos: {e}")
    import traceback
    traceback.print_exc()
    exit()

# =========================
# FILTRAR POR SEMANA SELECCIONADA (MIÉRCOLES AND JUEVES)
# =========================
if semana_analizar > 0:
    fecha_inicio, fecha_fin = obtener_rango_semana_miercoles_jueves(mes_analizar, anio_analizar, semana_analizar)
    mes_nombre = month_name[mes_analizar].upper()

    # Convertir fechas a formato date para comparación correcta
    fecha_inicio_date = fecha_inicio.date()
    fecha_fin_date = fecha_fin.date()

    # Filtrar por el rango de fechas de la semana seleccionada (miércoles y jueves)
    df_actual = df_filtrado[
        (df_filtrado[col_fecha].dt.date == fecha_inicio_date) | 
        (df_filtrado[col_fecha].dt.date == fecha_fin_date)
    ].copy()
else:
    # ANALIZAR TODOS LOS DATOS DISPONIBLES
    df_actual = df_filtrado.copy()
    if not df_actual.empty:
        fecha_inicio = df_actual[col_fecha].min()
        fecha_fin = df_actual[col_fecha].max()
        # Usar el mes predominante de los datos
        mes_analizar = df_actual[col_fecha].dt.month.mode()[0]
        anio_analizar = df_actual[col_fecha].dt.year.mode()[0]
        mes_nombre = month_name[mes_analizar].upper()
    else:
        print("⚠ No hay datos para procesar")
        exit()

if df_actual.empty:
    print(f"⚠ No hay datos para la semana {semana_analizar}: {fecha_inicio.strftime('%d/%m/%Y')} (mié) - {fecha_fin.strftime('%d/%m/%Y')} (jue)")
    print(f"📅 Fechas disponibles en los datos: {df_filtrado[col_fecha].min().strftime('%d/%m/%Y')} - {df_filtrado[col_fecha].max().strftime('%d/%m/%Y')}")
    exit()

print(f"✅ Datos para semana {semana_analizar}: {len(df_actual)} filas")
print(f"📅 Rango de fechas en datos filtrados: {df_actual[col_fecha].min().strftime('%d/%m/%Y')} - {df_actual[col_fecha].max().strftime('%d/%m/%Y')}")

# =========================
# SEPARAR EN SUCURSALES Y FRANQUICIAS
# =========================
# Filtrar sucursales
df_sucursales = df_actual[df_actual[col_tienda].isin(SUCURSALES)].copy()
df_franquicias = df_actual[~df_actual[col_tienda].isin(SUCURSALES)].copy()

print(f"📊 Sucursales: {len(df_sucursales)} filas, {df_sucursales[col_tienda].nunique()} tiendas")
print(f"📊 Franquicias: {len(df_franquicias)} filas, {df_franquicias[col_tienda].nunique()} tiendas")

# =========================
# FUNCIÓN PARA CREAR HOJA DE EXCEL
# =========================
def crear_hoja_excel(ws, df_data, titulo_hoja, tipo):
    """
    Crea una hoja de Excel con el formato profesional
    """
    # -------------------------
    # ESTILOS
    # -------------------------
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'),
                          top=Side(style='medium'), bottom=Side(style='medium'))
    
    if tipo == "sucursal":
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")  # Azul
        total_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")   # Verde
    else:
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")  # Púrpura
        total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")   # Naranja
    
    title_font = Font(bold=True, size=16, color="FFFFFF")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    total_font = Font(bold=True, size=11, color="000000")
    normal_font = Font(size=11)
    
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # -------------------------
    # RESUMEN POR TIENDA
    # -------------------------
    resumen = df_data.groupby(col_tienda).agg({
        col_nro_trx: pd.Series.nunique,
        'Beneficio_limpio': "sum"
    }).reset_index()
    
    # Calcular venta neta total (descuento del 10%)
    resumen["Total_Vendido"] = resumen['Beneficio_limpio'] / 0.10
    resumen.rename(columns={
        col_nro_trx: "Transacciones totales",
        'Beneficio_limpio': "Total Descuento",
        "Total_Vendido": "Venta Neta Total"
    }, inplace=True)
    
    # -------------------------
    # ENCABEZADO DEL REPORTE
    # -------------------------
    semanas_texto = ["ANTES PRIMERA", "PRIMERA", "SEGUNDA", "TERCERA", "CUARTA", "QUINTA"]
    
    ws.merge_cells('A1:D1')
    ws['A1'] = f"{titulo_hoja} - {semanas_texto[semana_analizar]} SEMANA DE {mes_nombre} {anio_analizar}"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws['A1'].fill = header_fill
    
    # Información resumen
    ws.merge_cells('A2:D2')
    tipo_texto = "Sucursales" if tipo == "sucursal" else "Franquicias"
    ws['A2'] = f"{tipo_texto} - Total Tiendas: {len(resumen)} | Transacciones: {resumen['Transacciones totales'].sum():,} | Período: {fecha_inicio.strftime('%d/%m/%Y')} (mié) - {fecha_fin.strftime('%d/%m/%Y')} (jue)"
    ws['A2'].font = Font(bold=True, size=10, color="000000")
    ws['A2'].alignment = center_align
    ws['A2'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Espacio
    ws.append([])
    
    # -------------------------
    # ENCABEZADOS DE COLUMNAS
    # -------------------------
    columnas_excel = ["Tienda", "Transacciones totales", "Total Descuento", "Venta Neta Total"]
    ws.append(columnas_excel)
    
    # Formato encabezados de columnas
    for col in range(1, 5):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border
    
    # -------------------------
    # DATOS POR TIENDA
    # -------------------------
    fila_inicio = 5
    for idx, row in resumen.iterrows():
        ws.cell(row=fila_inicio, column=1, value=row["Tienda"]).alignment = left_align
        ws.cell(row=fila_inicio, column=2, value=int(row["Transacciones totales"])).alignment = center_align
        ws.cell(row=fila_inicio, column=3, value=row["Total Descuento"])
        ws.cell(row=fila_inicio, column=4, value=row["Venta Neta Total"])
        fila_inicio += 1
    
    # Aplicar formato a todas las celdas de datos
    for row in ws.iter_rows(min_row=5, max_row=fila_inicio-1, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
            cell.font = normal_font
            if cell.column in [3, 4]:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = right_align
            elif cell.column == 2:
                cell.alignment = center_align
    
    # -------------------------
    # FILA DE TOTALES
    # =========================
    ws.cell(row=fila_inicio, column=1, value="TOTAL GENERAL").font = total_font
    ws.cell(row=fila_inicio, column=2, value=int(resumen["Transacciones totales"].sum())).font = total_font
    ws.cell(row=fila_inicio, column=3, value=resumen["Total Descuento"].sum()).font = total_font
    ws.cell(row=fila_inicio, column=4, value=resumen["Venta Neta Total"].sum()).font = total_font
    
    # Formato fila totales
    for col in range(1, 5):
        cell = ws.cell(row=fila_inicio, column=col)
        cell.fill = total_fill
        cell.border = thick_border
        cell.alignment = center_align
        if col in [3, 4]:
            cell.number_format = '"$"#,##0.00'
    
    # -------------------------
    # ESTADÍSTICAS ADICIONALES
    # -------------------------
    fila_inicio += 2
    
    ws.merge_cells(f'A{fila_inicio}:D{fila_inicio}')
    ws[f'A{fila_inicio}'] = f"ESTADÍSTICAS {tipo_texto.upper()}"
    ws[f'A{fila_inicio}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{fila_inicio}'].alignment = center_align
    ws[f'A{fila_inicio}'].fill = header_fill
    ws[f'A{fila_inicio}'].border = thin_border
    
    fila_inicio += 1
    ws[f'A{fila_inicio}'] = "Promedio de transacciones por tienda:"
    ws[f'B{fila_inicio}'] = resumen["Transacciones totales"].mean()
    ws[f'A{fila_inicio}'].font = Font(bold=True)
    ws[f'B{fila_inicio}'].font = Font(bold=True)
    ws[f'B{fila_inicio}'].alignment = center_align
    
    fila_inicio += 1
    ws[f'A{fila_inicio}'] = "Tienda con más transacciones:"
    ws[f'B{fila_inicio}'] = resumen.loc[resumen["Transacciones totales"].idxmax(), "Tienda"]
    ws[f'C{fila_inicio}'] = resumen["Transacciones totales"].max()
    ws[f'A{fila_inicio}'].font = Font(bold=True)
    ws[f'C{fila_inicio}'].font = Font(bold=True)
    ws[f'C{fila_inicio}'].alignment = center_align
    
    fila_inicio += 1
    ws[f'A{fila_inicio}'] = "Tienda con mayor venta neta:"
    ws[f'B{fila_inicio}'] = resumen.loc[resumen["Venta Neta Total"].idxmax(), "Tienda"]
    ws[f'C{fila_inicio}'] = resumen["Venta Neta Total"].max()
    ws[f'A{fila_inicio}'].font = Font(bold=True)
    ws[f'C{fila_inicio}'].font = Font(bold=True)
    ws[f'C{fila_inicio}'].number_format = '"$"#,##0.00'
    ws[f'C{fila_inicio}'].alignment = right_align
    
    # -------------------------
    # AJUSTES FINALES
    # -------------------------
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    
    ws.freeze_panes = 'A5'
    if len(resumen) > 0:
        ws.auto_filter.ref = f"A4:D{len(resumen) + 4}"
    
    return resumen

# =========================
# CREAR WORKBOOK CON MÚLTIPLES HOJAS
# =========================
wb = Workbook()

# Eliminar hoja por defecto si existe
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# Crear hoja de Sucursales
ws_sucursales = wb.create_sheet("Sucursales")
resumen_sucursales = crear_hoja_excel(ws_sucursales, df_sucursales, "INFORME VECINOS - SUCURSALES", "sucursal")

# Crear hoja de Franquicias
ws_franquicias = wb.create_sheet("Franquicias")
resumen_franquicias = crear_hoja_excel(ws_franquicias, df_franquicias, "INFORME VECINOS - FRANQUICIAS", "franquicia")

# Crear hoja de Resumen General
ws_resumen = wb.create_sheet("Resumen General")

# Configurar hoja de resumen
ws_resumen.merge_cells('A1:D1')
ws_resumen['A1'] = "RESUMEN GENERAL - COMPARATIVO"
ws_resumen['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws_resumen['A1'].alignment = Alignment(horizontal="center")
ws_resumen['A1'].fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

# Datos comparativos
data_comparativa = [
    ["", "Sucursales", "Franquicias", "Total"],
    ["N° de Tiendas", len(resumen_sucursales), len(resumen_franquicias), len(resumen_sucursales) + len(resumen_franquicias)],
    ["Total Transacciones", resumen_sucursales["Transacciones totales"].sum(), resumen_franquicias["Transacciones totales"].sum(), resumen_sucursales["Transacciones totales"].sum() + resumen_franquicias["Transacciones totales"].sum()],
    ["Total Descuento", resumen_sucursales["Total Descuento"].sum(), resumen_franquicias["Total Descuento"].sum(), resumen_sucursales["Total Descuento"].sum() + resumen_franquicias["Total Descuento"].sum()],
    ["Venta Neta Total", resumen_sucursales["Venta Neta Total"].sum(), resumen_franquicias["Venta Neta Total"].sum(), resumen_sucursales["Venta Neta Total"].sum() + resumen_franquicias["Venta Neta Total"].sum()]
]

for i, row in enumerate(data_comparativa, start=3):
    for j, value in enumerate(row, start=1):
        cell = ws_resumen.cell(row=i, column=j, value=value)
        if i == 3:  # Encabezados
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        else:
            if j > 1:  # Valores numéricos
                if i == 4:  # Fila 4: N° de Tiendas (formato número)
                    cell.number_format = '#,##0'
                elif i == 5:  # Fila 5: Total Transacciones (formato número)
                    cell.number_format = '#,##0'
                elif i == 6:  # Fila 6: Total Descuento (formato moneda)
                    cell.number_format = '"$"#,##0.00'
                elif i == 7:  # Fila 7: Venta Neta Total (formato moneda)
                    cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.font = Font(bold=True)

# Ajustar columnas
ws_resumen.column_dimensions['A'].width = 20
ws_resumen.column_dimensions['B'].width = 15
ws_resumen.column_dimensions['C'].width = 15
ws_resumen.column_dimensions['D'].width = 15

# =========================
# GUARDAR ARCHIVO EN GOOGLE DRIVE
# =========================
# Crear la carpeta de destino si no existe
if not os.path.exists(RUTA_DESTINO):
    try:
        os.makedirs(RUTA_DESTINO)
        print(f"✅ Carpeta creada: {RUTA_DESTINO}")
    except Exception as e:
        print(f"❌ No se pudo crear la carpeta de destino: {e}")
        # Guardar en directorio actual como respaldo
        RUTA_DESTINO = "."

nombre_archivo = f"Informe_Vecinos_Semana_{semana_analizar}_{mes_nombre}_{anio_analizar}.xlsx"
ruta_completa = os.path.join(RUTA_DESTINO, nombre_archivo)

try:
    wb.save(ruta_completa)
    print(f"✅ Informe profesional generado: {ruta_completa}")
except Exception as e:
    print(f"❌ Error al guardar en Google Drive: {e}")
    # Guardar en directorio actual como respaldo
    wb.save(nombre_archivo)
    print(f"✅ Informe guardado en directorio actual: {nombre_archivo}")

print(f"📊 Sucursales: {len(resumen_sucursales)} tiendas, {resumen_sucursales['Transacciones totales'].sum():,} transacciones")
print(f"📊 Franquicias: {len(resumen_franquicias)} tiendas, {resumen_franquicias['Transacciones totales'].sum():,} transacciones")
print(f"💰 Total Descuento: ${resumen_sucursales['Total Descuento'].sum() + resumen_franquicias['Total Descuento'].sum():,.2f}")
print(f"💰 Venta Neta Total: ${resumen_sucursales['Venta Neta Total'].sum() + resumen_franquicias['Venta Neta Total'].sum():,.2f}")
