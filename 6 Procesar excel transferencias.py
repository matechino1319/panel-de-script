import pandas as pd
import openpyxl
from pathlib import Path
from datetime import datetime
import os
import time
from script_runtime import get_input_dir

class ExtractorConDiscriminacion:
    """Extrae datos y discrimina entre sucursales y franquicias"""
    
    def __init__(self, ruta_carpeta="."):
        self.ruta_carpeta = ruta_carpeta
        self.archivo_log = None
        self.crear_archivo_log()
        
        # Lista de sucursales (sin el prefijo LY, solo la parte identificable)
        self.sucursales = {
            "Ballofet",
            "Velez",
            "Alem",
            "Cuadro Benegas",
            "CENTRO",
            "Libertador",
            "Bowen",
            "Alvear",
            "Atuel Norte",
            "DEPOSITODANI",
            "Deposito Logistica",  # Busca esta palabra clave
            "Distribución"
        }
        
    def crear_archivo_log(self):
        """Crear archivo de log"""
        fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.archivo_log = os.path.join(self.ruta_carpeta, f"LOG_extraccion_discriminada_{fecha}.txt")
        
        with open(self.archivo_log, 'w', encoding='utf-8') as f:
            f.write("=" * 100 + "\n")
            f.write("EXTRACCIÓN CON DISCRIMINACIÓN SUCURSAL/FRANQUICIA\n")
            f.write(f"Fecha: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write("=" * 100 + "\n\n")
    
    def log(self, mensaje):
        """Escribir en log y pantalla"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        linea = f"[{timestamp}] {mensaje}"
        print(linea)
        with open(self.archivo_log, 'a', encoding='utf-8') as f:
            f.write(linea + "\n")
    
    def discriminar_tienda(self, tienda_destino):
        """Discriminar si es sucursal o franquicia buscando coincidencias parciales"""
        if tienda_destino is None:
            return "Desconocido"
        
        # Convertir a string y limpiar
        tienda_str = str(tienda_destino).strip().upper()
        
        # Buscar si alguna palabra clave de sucursal está contenida en el nombre
        for sucursal in self.sucursales:
            if sucursal.upper() in tienda_str:
                return "Sucursal"
        
        # Si no encontró coincidencia, es franquicia
        return "Franquicia"
    
    def ejecutar(self):
        """Ejecutar extracción"""
        try:
            inicio = time.time()
            
            # PASO 1: Encontrar archivo
            self.log("=" * 100)
            self.log("PASO 1: Buscando archivo original...")
            self.log("=" * 100)
            
            archivos = list(Path(self.ruta_carpeta).glob("*.xlsx")) + \
                       list(Path(self.ruta_carpeta).glob("*.xls"))
            
            archivos_originales = [f for f in archivos if 'REPARADO' not in f.name and 'LIMPIO' not in f.name and 'FINAL' not in f.name and 'DISCRIMINAD' not in f.name]
            
            if not archivos_originales:
                self.log("❌ No hay archivo original")
                return False
            
            archivo = max(archivos_originales, key=lambda f: f.stat().st_size)
            self.log(f"✅ Archivo: {archivo.name}")
            
            # PASO 2: Cargar Excel
            self.log("\n" + "=" * 100)
            self.log("PASO 2: Cargando Excel...")
            self.log("=" * 100)
            
            tiempo_carga = time.time()
            wb = openpyxl.load_workbook(archivo, data_only=True)
            ws = wb.active
            
            self.log(f"✅ Cargado en {time.time() - tiempo_carga:.1f}s")
            self.log(f"   Dimensiones: {ws.max_row} filas x {ws.max_column} columnas")
            
            # PASO 3: Mapeo de columnas
            self.log("\n" + "=" * 100)
            self.log("PASO 3: Definiendo mapeo de columnas...")
            self.log("=" * 100)
            
            mapeo = {
                'Estado': 1,           # Col A
                'Número': 3,           # Col C
                'Recepción': 7,        # Col G
                'Tipo': 9,             # Col I
                'Tienda Origen': 11,   # Col K
                'Tienda Destino': 15,  # Col O
                'Fecha creación': 19,  # Col S
                'Fecha entrega': 21,   # Col U
                'Monto total': 32      # Col 32
            }
            
            self.log("✅ Mapeo de columnas:")
            for nombre, col_num in mapeo.items():
                self.log(f"   {nombre}: Columna {col_num}")
            
            # PASO 4: Buscar inicio de datos
            self.log("\n" + "=" * 100)
            self.log("PASO 4: Buscando inicio de datos...")
            self.log("=" * 100)
            
            fila_inicio = None
            for fila_idx in range(1, min(100, ws.max_row + 1)):
                celda_a = ws.cell(row=fila_idx, column=1).value
                if celda_a and ('CERRADA' in str(celda_a).upper() or 'CONFIRMADA' in str(celda_a).upper() or 'CANCELADA' in str(celda_a).upper()):
                    fila_inicio = fila_idx
                    self.log(f"✅ Datos encontrados en fila {fila_idx}")
                    break
            
            if fila_inicio is None:
                self.log("❌ No se encontró inicio de datos")
                return False
            
            # PASO 5: Extraer datos
            self.log("\n" + "=" * 100)
            self.log("PASO 5: Extrayendo datos...")
            self.log("=" * 100)
            
            datos_extraidos = []
            
            for fila_idx in range(fila_inicio, ws.max_row + 1):
                # Obtener valor de Estado para verificar si es fila de transferencia
                estado = ws.cell(row=fila_idx, column=mapeo['Estado']).value
                
                # Solo procesar filas que tienen Estado válido
                if estado and str(estado).upper() in ['CERRADA', 'CONFIRMADA', 'CANCELADA']:
                    fila_datos = {}
                    
                    for nombre_col, num_col in mapeo.items():
                        valor = ws.cell(row=fila_idx, column=num_col).value
                        
                        # Limpiar valor
                        if pd.isna(valor):
                            valor = None
                        elif isinstance(valor, str):
                            valor = valor.strip()
                        
                        fila_datos[nombre_col] = valor
                    
                    datos_extraidos.append(fila_datos)
                
                if fila_idx % 5000 == 0:
                    self.log(f"   Procesadas {fila_idx} filas... ({len(datos_extraidos)} transferencias)")
            
            self.log(f"\n✅ Extraídas {len(datos_extraidos)} transferencias")
            
            # PASO 6: Crear DataFrame
            self.log("\n" + "=" * 100)
            self.log("PASO 6: Creando tabla...")
            self.log("=" * 100)
            
            df = pd.DataFrame(datos_extraidos)
            
            # Convertir Monto total a numérico
            if 'Monto total' in df.columns:
                df['Monto total'] = pd.to_numeric(df['Monto total'], errors='coerce')
            
            self.log(f"✅ Tabla creada: {len(df)} filas x {len(df.columns)} columnas")
            
            # PASO 7: Agregar discriminación de sucursal/franquicia
            self.log("\n" + "=" * 100)
            self.log("PASO 7: Discriminando sucursales vs franquicias...")
            self.log("=" * 100)
            
            # Primero mostrar tiendas únicas para debuggeo
            self.log(f"\n   Tiendas destino únicas encontradas:")
            tiendas_unicas = df['Tienda Destino'].unique()
            for i, tienda in enumerate(tiendas_unicas[:20], 1):  # Mostrar primeras 20
                self.log(f"      {i}. {tienda}")
            if len(tiendas_unicas) > 20:
                self.log(f"      ... y {len(tiendas_unicas) - 20} más")
            
            # Aplicar discriminación
            df['Tipo Tienda'] = df['Tienda Destino'].apply(self.discriminar_tienda)
            
            sucursales_count = (df['Tipo Tienda'] == 'Sucursal').sum()
            franquicias_count = (df['Tipo Tienda'] == 'Franquicia').sum()
            desconocidas_count = (df['Tipo Tienda'] == 'Desconocido').sum()
            
            self.log(f"\n✅ Discriminación completada:")
            self.log(f"   Sucursales: {sucursales_count}")
            self.log(f"   Franquicias: {franquicias_count}")
            self.log(f"   Desconocidas: {desconocidas_count}")
            
            # PASO 8: Mostrar vista previa
            self.log("\n" + "=" * 100)
            self.log("PASO 8: Vista previa de datos...")
            self.log("=" * 100)
            
            self.log(f"\n📋 Columnas:")
            for i, col in enumerate(df.columns, 1):
                self.log(f"   {i}. {col}")
            
            self.log(f"\n📊 Primeras 5 filas:")
            for idx in range(min(5, len(df))):
                self.log(f"\n   Fila {idx + 1}:")
                for col in df.columns:
                    valor = df.iloc[idx][col]
                    if pd.notna(valor):
                        self.log(f"      {col}: {valor}")
            
            # Estadísticas
            self.log(f"\n📈 Estadísticas:")
            self.log(f"   Estados únicos: {list(df['Estado'].unique())}")
            self.log(f"   Tiendas destino únicas: {len(df['Tienda Destino'].unique())}")
            self.log(f"   Monto total: ${df['Monto total'].sum():,.2f}")
            
            # Desglose por tipo de tienda
            self.log(f"\n📊 Desglose por tipo de tienda:")
            for tipo in ['Sucursal', 'Franquicia', 'Desconocido']:
                df_tipo = df[df['Tipo Tienda'] == tipo]
                if len(df_tipo) > 0:
                    monto_tipo = df_tipo['Monto total'].sum()
                    self.log(f"   {tipo}: {len(df_tipo)} transferencias | Monto: ${monto_tipo:,.2f}")
            
            # PASO 9: Guardar
            self.log("\n" + "=" * 100)
            self.log("PASO 9: Guardando archivo...")
            self.log("=" * 100)
            
            fecha = datetime.now().strftime("%Y%m%d_%H%M%S")
            nombre = f"transferencias_DISCRIMINADAS_{fecha}.xlsx"
            ruta_salida = os.path.join(self.ruta_carpeta, nombre)
            
            df.to_excel(ruta_salida, index=False, sheet_name='Transferencias')
            
            tiempo_total = time.time() - inicio
            
            self.log(f"\n✅ Guardado: {nombre}")
            self.log(f"⏱️  Tiempo total: {tiempo_total:.1f}s")
            
            self.log("\n" + "=" * 100)
            self.log("✅ EXTRACCIÓN Y DISCRIMINACIÓN COMPLETADA CORRECTAMENTE")
            self.log("=" * 100)
            self.log(f"\n📊 RESUMEN:")
            self.log(f"   Archivo guardado: {ruta_salida}")
            self.log(f"   Transferencias extraídas: {len(df)}")
            self.log(f"   Columnas: {len(df.columns)}")
            self.log(f"   Sucursales: {sucursales_count}")
            self.log(f"   Franquicias: {franquicias_count}")
            
            return True
            
        except Exception as e:
            self.log(f"\n❌ ERROR: {e}")
            import traceback
            self.log(f"\nDetalles:\n{traceback.format_exc()}")
            return False


if __name__ == "__main__":
    import sys
    # Configurar salida UTF-8 para evitar errores con emojis en Windows
    if sys.stdout.encoding != 'utf-8':
        try:
            sys.stdout.reconfigure(encoding='utf-8')
        except AttributeError:
            # Fallback para versiones de python < 3.7
            import codecs
            sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())

    try:
        # Obtener la ruta donde está ubicado el script
        ruta = get_input_dir()

        print(f"📁 Carpeta del script: {ruta}\n")

        extractor = ExtractorConDiscriminacion(ruta_carpeta=ruta)
        resultado = extractor.ejecutar()

        if resultado:
            print(f"\n✅ Extracción y discriminación completada")
            print(f"   Log: {extractor.archivo_log}")

    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()

    finally:
        print("\n" + "=" * 100)
        print("Proceso finalizado.")
        print("=" * 100)
