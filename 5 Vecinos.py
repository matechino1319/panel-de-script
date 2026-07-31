#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Informe Vecinos - Versión totalmente automática (lee .xlsx/.xls/.csv sin preguntar,
# no pide inputs y NO filtra por fechas: analiza todo el archivo y genera el informe)
# Autor: Adaptado para Gonzapereyraa
# Fecha: 2025-11-03

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from calendar import month_name
from datetime import datetime, timedelta
import re
import unicodedata
import os
import glob
import sys
from script_runtime import get_input_dir, get_output_dir, get_input_file

# Configurar salida UTF-8 para evitar errores con emojis en Windows
if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except AttributeError:
        # Fallback para versiones de python < 3.7
        import codecs
        sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())

# -------------------------
# CONFIGURACIÓN
# -------------------------
# Carpeta de salida (intenta guardar en Drive, si falla guarda en carpeta actual)
RUTA_DESTINO = get_output_dir(r"G:\Mi unidad\ARCHIVOS_COMPARTIDOS_LAYUNTA\SISTEMAS\INFORMES VECINOS")

# Lista de sucursales (las demás serán consideradas franquicias)
SUCURSALES = [
    "LY01-Ballofet",
    "LY02-Velez",
    "LY04-Alem",
    "LY07-Cuadro Benegas",
    "LY19-Alvear",
    "LY22-CENTRO",
    "LY25-Montoya",
    "LY25-Alberdi",
    "LY34-Bowen",
    "LY37-Libertador",
    "LY42-Atuel Norte"
]
# -------------------------
def encontrar_archivo_entrada():
    archivo_directo = get_input_file()
    if archivo_directo:
        return archivo_directo

    patrones_excel = ["*.xlsx", "*.xlsm", "*.xls"]
    archivos = []
    for p in patrones_excel:
        archivos.extend(glob.glob(os.path.join(get_input_dir(), p)))
    archivos = sorted(list(dict.fromkeys(archivos)))  # únicos y ordenados

    if archivos:
        return archivos[0]  # primer Excel encontrado

    # Si no hay Excel, buscar CSV
    archivos_csv = sorted(glob.glob(os.path.join(get_input_dir(), "*.csv")))
    if archivos_csv:
        return archivos_csv[0]

    return None

archivo_entrada = encontrar_archivo_entrada()
if archivo_entrada is None:
    print("❌ No se encontraron archivos .xlsx/.xls/.csv en el directorio actual:", os.getcwd())
    sys.exit(1)

print(f"✅ Archivo seleccionado automáticamente: {archivo_entrada}")

# -------------------------
# LECTURA DEL ARCHIVO (Excel o CSV)
# -------------------------
ext = os.path.splitext(archivo_entrada)[1].lower()
try:
    if ext in ['.xlsx', '.xls', '.xlsm']:
        # lee la primera hoja por defecto
        df = pd.read_excel(archivo_entrada, sheet_name=0, engine='openpyxl')
    elif ext == '.csv':
        # Intentar con ; y luego con ,
        try:
            df = pd.read_csv(archivo_entrada, encoding="utf-8", sep=';', on_bad_lines='skip')
        except Exception:
            df = pd.read_csv(archivo_entrada, encoding="latin-1", sep=';', on_bad_lines='skip')
    else:
        raise ValueError("Formato de archivo no soportado")
    print(f"✅ Archivo leído. Filas: {len(df)}, Columnas: {len(df.columns)}")
except Exception as e:
    print("❌ Error al leer el archivo:", e)
    sys.exit(1)

print("Columnas originales:", df.columns.tolist())

# -------------------------
# IDENTIFICACIÓN Y NORMALIZACIÓN DE COLUMNAS
# -------------------------
def normalizar_nombre(col):
    coln = unicodedata.normalize('NFKD', str(col))
    coln = coln.encode('ascii', errors='ignore').decode('ascii')
    coln = re.sub(r'[^a-zA-Z0-9]', '_', coln)
    coln = re.sub(r'_+', '_', coln)
    return coln.strip('_')

original_columns = df.columns.tolist()
new_columns = []
for col in original_columns:
    col_str = str(col)
    if col_str == '$Beneficio':
        col_normalized = 'Beneficio_monetario'
    elif col_str == '%Beneficio':
        col_normalized = 'Beneficio_porcentaje'
    elif col_str == 'Beneficio':
        col_normalized = 'Beneficio_general'
    else:
        col_normalized = normalizar_nombre(col_str)
    base_name = col_normalized
    counter = 1
    while col_normalized in new_columns:
        col_normalized = f"{base_name}_{counter}"
        counter += 1
    new_columns.append(col_normalized)
df.columns = new_columns
print("Columnas normalizadas:", df.columns.tolist())

def encontrar_columna(patrones):
    for col in df.columns:
        col_lower = col.lower()
        for patron in patrones:
            if patron in col_lower:
                return col
    return None

# Detectar columnas clave (con fallbacks)
col_beneficio = None
prioritarias = ['beneficio_monetario', '$beneficio', 'dbeneficio', 'beneficio_$', 'monto_beneficio']
for col in df.columns:
    cl = col.lower()
    for p in prioritarias:
        if p in cl:
            col_beneficio = col
            break
    if col_beneficio:
        break
if col_beneficio is None:
    for col in df.columns:
        if 'beneficio' in col.lower() and 'porcentaje' not in col.lower():
            col_beneficio = col
            break
if col_beneficio is None:
    col_beneficio = encontrar_columna(['monto', 'valor', 'importe', 'descuento', 'discount'])

col_mensaje = encontrar_columna(['mensaje', 'promocion', 'promocion', 'promoción', 'promo'])
col_nro_trx = encontrar_columna(['nro', 'trx', 'transaccion', 'numero', 'nro_trx'])
col_fecha = encontrar_columna(['fecha', 'inicio', 'date', 'fecha_inicio_trx'])
col_tienda = encontrar_columna(['tienda', 'sucursal', 'store', 'local'])

# Fallbacks rápidos
if col_nro_trx is None:
    for col in df.columns:
        if 'nro' in col.lower() and 'fecha' not in col.lower():
            col_nro_trx = col
            break
if col_fecha is None:
    for col in df.columns:
        if 'fecha' in col.lower():
            col_fecha = col
            break

columnas_requeridas = {
    'Mensaje': col_mensaje,
    'NroTrx': col_nro_trx,
    'Fecha': col_fecha,
    'Beneficio': col_beneficio,
    'Tienda': col_tienda
}

faltantes = [k for k,v in columnas_requeridas.items() if v is None]
if faltantes:
    print("❌ No se encontraron columnas requeridas:", faltantes)
    print("Columnas disponibles:", df.columns.tolist())
    sys.exit(1)

print("🔍 Columnas identificadas:", columnas_requeridas)

# -------------------------
# FILTRAR PROMOCIONES (vecinos) - patrón flexible
# -------------------------
mensajes_filtrar = [
    "10% de descuento a VECINOS",
    "10% de descuento a VECINOS DEBITO",
    "JUEVES 10% de descuento a VECINOS",
    "Miercoles 10% de descuento a VECINOS",
    "10% vecinos",
    "descuento a vecinos",
    "descuento vecinos"
]
patron_mensajes = "|".join([re.escape(m).lower() for m in mensajes_filtrar])

# Convertir a str y filtrar - si columna mensaje contiene NaN, astype(str) lo manejará
df_filtrado = df[df[col_mensaje].astype(str).str.lower().str.contains(patron_mensajes, na=False)].copy()
if df_filtrado.empty:
    print("⚠ No se encontraron filas que coincidan con las promociones buscadas.")
    sys.exit(0)

print(f"✅ Filtrado promociones: {len(df_filtrado)} filas")

# -------------------------
# LIMPIEZA Y CONVERSIÓN MONETARIA
# -------------------------
def limpiar_valor_monetario(valor):
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    s = str(valor).strip()
    if s == '' or s.lower() in ['nan', 'none']:
        return 0.0
    s = re.sub(r'[^\d,-.]', '', s)
    if ',' in s and '.' in s:
        if s.find(',') > s.find('.'):
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '')
    else:
        s = s.replace(',', '.')
    if s.count('.') > 1:
        parts = s.split('.')
        s = ''.join(parts[:-1]) + '.' + parts[-1]
    try:
        return abs(float(s))
    except Exception:
        return 0.0

df_filtrado['Beneficio_limpio'] = df_filtrado[col_beneficio].apply(limpiar_valor_monetario)

# -------------------------
# NOTA: NO SE FILTRA POR FECHAS - se analiza todo el dataset filtrado por mensajes
# -------------------------
df_actual = df_filtrado.copy()
print(f"✅ Analizando todo el dataset filtrado (sin filtrar por fecha): {len(df_actual)} filas")

# -------------------------
# Separar sucursales y franquicias
# -------------------------
df_sucursales = df_actual[df_actual[col_tienda].isin(SUCURSALES)].copy()
df_franquicias = df_actual[~df_actual[col_tienda].isin(SUCURSALES)].copy()

print(f"📊 Sucursales: {len(df_sucursales)} filas, {df_sucursales[col_tienda].nunique()} tiendas")
print(f"📊 Franquicias: {len(df_franquicias)} filas, {df_franquicias[col_tienda].nunique()} tiendas")

# -------------------------
# FUNCION: CREAR HOJA DE EXCEL (formato profesional)
# -------------------------
def crear_hoja_excel(ws, df_data, titulo_hoja, tipo, periodo_text="PERIODO COMPLETO"):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'),
                          top=Side(style='medium'), bottom=Side(style='medium'))
    if tipo == "sucursal":
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        total_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    else:
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        total_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

    title_font = Font(bold=True, size=16, color="FFFFFF")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    total_font = Font(bold=True, size=11, color="000000")
    normal_font = Font(size=11)

    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    resumen = df_data.groupby(col_tienda).agg({
        col_nro_trx: pd.Series.nunique,
        'Beneficio_limpio': "sum"
    }).reset_index()

    resumen["Total_Vendido"] = resumen['Beneficio_limpio'] / 0.10
    resumen.rename(columns={
        col_nro_trx: "Transacciones totales",
        'Beneficio_limpio': "Total Descuento",
        "Total_Vendido": "Venta Neta Total"
    }, inplace=True)

    ws.merge_cells('A1:D1')
    ws['A1'] = f"{titulo_hoja} - {periodo_text}"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws['A1'].fill = header_fill

    ws.merge_cells('A2:D2')
    tipo_texto = "Sucursales" if tipo == "sucursal" else "Franquicias"
    ws['A2'] = f"{tipo_texto} - Total Tiendas: {len(resumen)} | Transacciones: {resumen['Transacciones totales'].sum():,} | Período: {periodo_text}"
    ws['A2'].font = Font(bold=True, size=10, color="000000")
    ws['A2'].alignment = center_align
    ws['A2'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

    ws.append([])
    columnas_excel = ["Tienda", "Transacciones totales", "Total Descuento", "Venta Neta Total"]
    ws.append(columnas_excel)

    for col_idx in range(1, 5):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border

    fila_inicio = 5
    for idx, row in resumen.iterrows():
        ws.cell(row=fila_inicio, column=1, value=row[col_tienda]).alignment = left_align
        ws.cell(row=fila_inicio, column=2, value=int(row["Transacciones totales"])).alignment = center_align
        ws.cell(row=fila_inicio, column=3, value=row["Total Descuento"])
        ws.cell(row=fila_inicio, column=4, value=row["Venta Neta Total"])
        fila_inicio += 1

    for row in ws.iter_rows(min_row=5, max_row=fila_inicio-1, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
            cell.font = normal_font
            if cell.column in [3, 4]:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = right_align
            elif cell.column == 2:
                cell.alignment = center_align

    ws.cell(row=fila_inicio, column=1, value="TOTAL GENERAL").font = total_font
    ws.cell(row=fila_inicio, column=2, value=int(resumen["Transacciones totales"].sum())).font = total_font
    ws.cell(row=fila_inicio, column=3, value=resumen["Total Descuento"].sum()).font = total_font
    ws.cell(row=fila_inicio, column=4, value=resumen["Venta Neta Total"].sum()).font = total_font

    for col_idx in range(1, 5):
        cell = ws.cell(row=fila_inicio, column=col_idx)
        cell.fill = total_fill
        cell.border = thick_border
        cell.alignment = center_align
        if col_idx in [3, 4]:
            cell.number_format = '"$"#,##0.00'

    fila_inicio += 2
    ws.merge_cells(f'A{fila_inicio}:D{fila_inicio}')
    ws[f'A{fila_inicio}'] = f"ESTADÍSTICAS {tipo_texto.upper()}"
    ws[f'A{fila_inicio}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{fila_inicio}'].alignment = center_align
    ws[f'A{fila_inicio}'].fill = header_fill
    ws[f'A{fila_inicio}'].border = thin_border

    fila_inicio += 1
    ws[f'A{fila_inicio}'] = "Promedio de transacciones por tienda:"
    ws[f'B{fila_inicio}'] = resumen["Transacciones totales"].mean() if len(resumen) > 0 else 0
    ws[f'A{fila_inicio}'].font = Font(bold=True)
    ws[f'B{fila_inicio}'].font = Font(bold=True)
    ws[f'B{fila_inicio}'].alignment = center_align

    fila_inicio += 1
    if len(resumen) > 0:
        ws[f'A{fila_inicio}'] = "Tienda con más transacciones:"
        ws[f'B{fila_inicio}'] = resumen.loc[resumen["Transacciones totales"].idxmax(), col_tienda]
        ws[f'C{fila_inicio}'] = resumen["Transacciones totales"].max()
        ws[f'A{fila_inicio}'].font = Font(bold=True)
        ws[f'C{fila_inicio}'].font = Font(bold=True)
        ws[f'C{fila_inicio}'].alignment = center_align

        fila_inicio += 1
        ws[f'A{fila_inicio}'] = "Tienda con mayor venta neta:"
        ws[f'B{fila_inicio}'] = resumen.loc[resumen["Venta Neta Total"].idxmax(), col_tienda]
        ws[f'C{fila_inicio}'] = resumen["Venta Neta Total"].max()
        ws[f'A{fila_inicio}'].font = Font(bold=True)
        ws[f'C{fila_inicio}'].font = Font(bold=True)
        ws[f'C{fila_inicio}'].number_format = '"$"#,##0.00'
        ws[f'C{fila_inicio}'].alignment = right_align

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18

    ws.freeze_panes = 'A5'
    if len(resumen) > 0:
        ws.auto_filter.ref = f"A4:D{len(resumen) + 4}"

    return resumen

# -------------------------
# GENERAR WORKBOOK Y HOJAS
# -------------------------
wb = Workbook()
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

periodo_text = "PERIODO COMPLETO (SIN FILTRADO POR FECHAS)"

ws_suc = wb.create_sheet("Sucursales")
res_suc = crear_hoja_excel(ws_suc, df_sucursales, "INFORME VECINOS - SUCURSALES", "sucursal", periodo_text)

ws_fran = wb.create_sheet("Franquicias")
res_fran = crear_hoja_excel(ws_fran, df_franquicias, "INFORME VECINOS - FRANQUICIAS", "franquicia", periodo_text)

ws_resumen = wb.create_sheet("Resumen General")
ws_resumen.merge_cells('A1:D1')
ws_resumen['A1'] = "RESUMEN GENERAL - COMPARATIVO (PERIODO COMPLETO)"
ws_resumen['A1'].font = Font(bold=True, size=16, color="FFFFFF")
ws_resumen['A1'].alignment = Alignment(horizontal="center")
ws_resumen['A1'].fill = PatternFill(start_color="0070C0", end_color="0070C0", fill_type="solid")

data_comparativa = [
    ["", "Sucursales", "Franquicias", "Total"],
    ["N° de Tiendas", len(res_suc), len(res_fran), len(res_suc) + len(res_fran)],
    ["Total Transacciones", res_suc["Transacciones totales"].sum(), res_fran["Transacciones totales"].sum(), res_suc["Transacciones totales"].sum() + res_fran["Transacciones totales"].sum()],
    ["Total Descuento", res_suc["Total Descuento"].sum(), res_fran["Total Descuento"].sum(), res_suc["Total Descuento"].sum() + res_fran["Total Descuento"].sum()],
    ["Venta Neta Total", res_suc["Venta Neta Total"].sum(), res_fran["Venta Neta Total"].sum(), res_suc["Venta Neta Total"].sum() + res_fran["Venta Neta Total"].sum()]
]

for i, row in enumerate(data_comparativa, start=3):
    for j, value in enumerate(row, start=1):
        cell = ws_resumen.cell(row=i, column=j, value=value)
        if i == 3:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        else:
            if j > 1:
                if i == 4:
                    cell.number_format = '#,##0'
                elif i == 5:
                    cell.number_format = '#,##0'
                elif i == 6:
                    cell.number_format = '"$"#,##0.00'
                elif i == 7:
                    cell.number_format = '"$"#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.font = Font(bold=True)

ws_resumen.column_dimensions['A'].width = 20
ws_resumen.column_dimensions['B'].width = 15
ws_resumen.column_dimensions['C'].width = 15
ws_resumen.column_dimensions['D'].width = 15

# -------------------------
# GUARDAR ARCHIVO
# -------------------------
if not os.path.exists(RUTA_DESTINO):
    try:
        os.makedirs(RUTA_DESTINO)
        print(f"✅ Carpeta creada: {RUTA_DESTINO}")
    except Exception as e:
        print("⚠ No se pudo crear la carpeta de destino, se guardará en el directorio actual:", e)
        RUTA_DESTINO = "."

# Nombre del archivo con timestamp corto
timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
nombre_archivo = f"Informe_Vecinos_{timestamp}.xlsx"
ruta_completa = os.path.join(RUTA_DESTINO, nombre_archivo)

try:
    wb.save(ruta_completa)
    print(f"✅ Informe generado: {ruta_completa}")
except Exception as e:
    print("❌ Error al guardar en la ruta destino:", e)
    try:
        wb.save(nombre_archivo)
        print("✅ Informe guardado en directorio actual:", nombre_archivo)
    except Exception as e2:
        print("❌ No se pudo guardar el informe:", e2)
        sys.exit(1)

# -------------------------
# RESUMEN POR CONSOLA
# -------------------------
total_descuento = res_suc["Total Descuento"].sum() + res_fran["Total Descuento"].sum()
total_venta_neta = res_suc["Venta Neta Total"].sum() + res_fran["Venta Neta Total"].sum()
trans_suc = res_suc["Transacciones totales"].sum()
trans_fran = res_fran["Transacciones totales"].sum()

print(f"📊 Sucursales: {len(res_suc)} tiendas, {trans_suc:,} transacciones")
print(f"📊 Franquicias: {len(res_fran)} tiendas, {trans_fran:,} transacciones")
print(f"💰 Total Descuento: ${total_descuento:,.2f}")
print(f"💰 Venta Neta Total: ${total_venta_neta:,.2f}")

# Fin del script
