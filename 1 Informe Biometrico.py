"""
SCRIPT GENERADOR DE REPORTE BIOMÉTRICO - EXCEL ÚNICO CON TABLAS
Genera un Excel con 3 hojas en formato tabla:
1. Datos completos de asistencia
2. Resumen de horas por persona (sin total general)
3. Registros incompletos (falta entrada o salida)
"""

import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import sys
from typing import Tuple, Dict, List, Optional
import glob
import logging
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

# Configurar salida UTF-8 para evitar errores con emojis en Windows
if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except AttributeError:
        # Fallback para versiones de python < 3.7
        import codecs
        sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())

class ReporteBiometrico:
    """Generador de reportes biométricos en Excel con formato tabla"""
    
    def __init__(self, directorio_trabajo: str = "."):
        """Inicializa el generador de reportes"""
        self.directorio_trabajo = directorio_trabajo
        self.archivo_entrada = None
        self.df_original = None
        self.df_procesado = None
        self.directorio_salida = "reportes_biometricos"
        self.archivo_log = None
        self.logger = None
        self.archivo_excel_salida = None
        
        # Crear directorio de salida
        if not os.path.exists(self.directorio_salida):
            os.makedirs(self.directorio_salida)
        
        self._configurar_logging()
        self._detectar_archivo_excel()
    
    def _configurar_logging(self):
        """Configura el sistema de logging"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        self.archivo_log = f"{self.directorio_salida}/logs_reporte_{timestamp}.txt"
        
        self.logger = logging.getLogger('ReporteBiometrico')
        self.logger.setLevel(logging.DEBUG)
        self.logger.handlers = []
        
        # Handler para archivo
        file_handler = logging.FileHandler(self.archivo_log, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        
        # Handler para consola
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        
        # Formato
        formatter = logging.Formatter(
            '%(asctime)s - %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S'
        )
        file_handler.setFormatter(formatter)
        console_handler.setFormatter(formatter)
        
        self.logger.addHandler(file_handler)
        self.logger.addHandler(console_handler)
        
        self.logger.info("=" * 80)
        self.logger.info("GENERADOR DE REPORTE BIOMÉTRICO - EXCEL ÚNICO CON TABLAS")
        self.logger.info("=" * 80)
        self.logger.info(f"Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        self.logger.info(f"Usuario: {os.getenv('USERNAME', 'Usuario')}")
    
    def _detectar_archivo_excel(self) -> bool:
        """Detecta automáticamente el archivo Excel (ignorando temporales)"""
        self.logger.info(f"🔍 Buscando archivos Excel en: {self.directorio_trabajo}")
        
        patrones = ['*.xlsx', '*.xls', '*.XLSX', '*.XLS']
        archivos_encontrados = []
        
        for patron in patrones:
            ruta_patron = os.path.join(self.directorio_trabajo, patron)
            archivos = glob.glob(ruta_patron)
            
            # Filtrar archivos temporales
            archivos = [f for f in archivos if not os.path.basename(f).startswith('~$')]
            archivos_encontrados.extend(archivos)
        
        if not archivos_encontrados:
            self.logger.error("❌ No se encontraron archivos Excel válidos")
            print("❌ Error: No se encontraron archivos Excel válidos")
            return False
        
        # Usar el archivo más reciente
        archivo_seleccionado = max(archivos_encontrados, key=os.path.getctime)
        self.archivo_entrada = archivo_seleccionado
        
        self.logger.info(f"✅ Archivo Excel detectado: {os.path.basename(self.archivo_entrada)}")
        self.logger.info(f"   Tamaño: {os.path.getsize(self.archivo_entrada) / 1024:.2f} KB")
        self.logger.info(f"   Fecha: {datetime.fromtimestamp(os.path.getmtime(self.archivo_entrada))}")
        
        print(f"✅ Archivo Excel detectado: {os.path.basename(self.archivo_entrada)}")
        return True
    
    def cargar_datos(self) -> pd.DataFrame:
        """Carga los datos del archivo Excel"""
        if not self.archivo_entrada:
            self.logger.error("❌ No hay archivo Excel detectado")
            return None
        
        try:
            self.logger.info(f"📂 Cargando datos...")
            
            # Leer sin procesar fechas
            df_raw = pd.read_excel(self.archivo_entrada, header=None, dtype=str)
            
            self.logger.debug(f"   Filas totales: {len(df_raw)}")
            
            # Detectar fila de inicio de datos
            fila_datos = self._detectar_fila_datos_real(df_raw)
            self.logger.info(f"   Fila de inicio detectada: {fila_datos}")
            
            # Extraer datos
            df_datos = df_raw.iloc[fila_datos:].copy()
            df_datos = df_datos.reset_index(drop=True)
            
            # Mapear columnas
            df_limpio = pd.DataFrame({
                'Fecha': df_datos.iloc[:, 1],
                'Hora': df_datos.iloc[:, 2],
                'Sensor': df_datos.iloc[:, 3],
                'Tipo Persona': df_datos.iloc[:, 4],
                'Nombre': df_datos.iloc[:, 7]
            })
            
            # Limpiar datos
            df_limpio = df_limpio.dropna(how='all')
            df_limpio = df_limpio[df_limpio['Nombre'].notna() & 
                                  (df_limpio['Nombre'].astype(str).str.strip() != '') &
                                  (df_limpio['Nombre'].astype(str).str.lower() != 'todos') &
                                  (df_limpio['Fecha'].notna()) &
                                  (df_limpio['Fecha'].astype(str).str.strip() != '')]
            
            # Limpiar espacios
            for col in df_limpio.columns:
                df_limpio[col] = df_limpio[col].astype(str).str.strip()
            
            self.df_original = df_limpio.reset_index(drop=True)
            
            self.logger.info(f"✅ Datos cargados: {len(self.df_original)} registros")
            print(f"✅ Datos cargados: {len(self.df_original)} registros")
            
            return self.df_original
            
        except Exception as e:
            self.logger.error(f"❌ Error al cargar: {str(e)}", exc_info=True)
            return None
    
    def _detectar_fila_datos_real(self, df: pd.DataFrame) -> int:
        """Detecta la fila donde comienzan los datos reales"""
        try:
            for idx, row in df.iterrows():
                for cell in row:
                    if pd.notna(cell):
                        cell_str = str(cell).strip()
                        if self._es_fecha_valida(cell_str):
                            self.logger.debug(f"   Primera fecha en fila {idx}: {cell_str}")
                            return idx
            return 6
        except:
            return 6
    
    def _es_fecha_valida(self, texto: str) -> bool:
        """Verifica si es una fecha válida DD/MM/YY"""
        try:
            if len(texto) == 8 and texto.count('/') == 2:
                partes = texto.split('/')
                if len(partes) == 3:
                    dia = int(partes[0])
                    mes = int(partes[1])
                    año = int(partes[2])
                    return (1 <= dia <= 31) and (1 <= mes <= 12) and (0 <= año <= 99)
            return False
        except:
            return False
    
    def procesar_datos(self) -> pd.DataFrame:
        """Procesa y valida los datos"""
        try:
            self.logger.info("🔄 Procesando datos...")
            
            df = self.df_original.copy()
            
            # Validar tipo
            df['Tipo Persona'] = df['Tipo Persona'].astype(str).str.lower().str.strip()
            df = df[df['Tipo Persona'].isin(['entrada', 'salida'])]
            
            # Validar hora
            df = df[df['Hora'].notna() & (df['Hora'].astype(str).str.strip() != '')]
            
            self.logger.info(f"   ✅ Datos procesados: {len(df)} registros válidos")
            
            self.df_procesado = df.reset_index(drop=True)
            return df
            
        except Exception as e:
            self.logger.error(f"❌ Error procesando: {str(e)}", exc_info=True)
            return None
    
    def generar_datos_completos(self) -> pd.DataFrame:
        """Genera la primera hoja: Datos completos"""
        try:
            self.logger.info("📄 Generando datos completos...")
            
            df = self.df_procesado.copy()
            
            # Ordenar
            df_completos = df[['Fecha', 'Hora', 'Sensor', 'Tipo Persona', 'Nombre']].copy()
            df_completos = df_completos.sort_values(['Nombre', 'Fecha', 'Hora']).reset_index(drop=True)
            
            self.logger.info(f"   ✅ {len(df_completos)} registros en datos completos")
            
            return df_completos
            
        except Exception as e:
            self.logger.error(f"❌ Error generando datos completos: {str(e)}", exc_info=True)
            return pd.DataFrame()
    
    def generar_resumen_horas(self) -> pd.DataFrame:
        """Genera la segunda hoja: Resumen de horas por persona CON TOTAL ACUMULADO"""
        try:
            self.logger.info("📊 Generando resumen de horas con total acumulado...")
            
            df = self.df_procesado.copy()
            
            resumen_detalle = []
            
            # Agrupar por persona y fecha
            for (nombre, fecha), grupo in df.groupby(['Nombre', 'Fecha']):
                
                # Obtener entradas y salidas
                entradas = grupo[grupo['Tipo Persona'] == 'entrada']['Hora'].values
                salidas = grupo[grupo['Tipo Persona'] == 'salida']['Hora'].values
                
                # Si tiene entrada y salida
                if len(entradas) > 0 and len(salidas) > 0:
                    hora_entrada = str(entradas[0]).strip()
                    hora_salida = str(salidas[-1]).strip()
                    
                    horas = self._calcular_horas(hora_entrada, hora_salida)
                    
                    if horas > 0:
                        resumen_detalle.append({
                            'Nombre': nombre,
                            'Fecha': fecha,
                            'Entrada': hora_entrada,
                            'Salida': hora_salida,
                            'Horas del Día': round(horas, 2)
                        })
            
            # Crear DataFrame detallado
            df_resumen_detalle = pd.DataFrame(resumen_detalle)
            
            if len(df_resumen_detalle) > 0:
                # Ordenar por persona y fecha
                df_resumen_detalle = df_resumen_detalle.sort_values(['Nombre', 'Fecha']).reset_index(drop=True)
                
                # Calcular total acumulado por persona
                df_resumen_final = df_resumen_detalle.groupby('Nombre', as_index=False).agg({
                    'Horas del Día': 'sum',
                    'Fecha': 'count'
                }).rename(columns={'Fecha': 'Días Registrados'})
                
                df_resumen_final = df_resumen_final.rename(columns={'Horas del Día': 'Total Horas Trabajadas'})
                df_resumen_final['Promedio Diario'] = (df_resumen_final['Total Horas Trabajadas'] / df_resumen_final['Días Registrados']).round(2)
                
                # Reordenar columnas
                df_resumen_final = df_resumen_final[['Nombre', 'Total Horas Trabajadas', 'Días Registrados', 'Promedio Diario']]
                df_resumen_final = df_resumen_final.sort_values('Total Horas Trabajadas', ascending=False)
                
                # SIN fila de totales
                df_resumen_final = df_resumen_final.reset_index(drop=True)
                
                self.logger.info(f"   ✅ Resumen generado: {len(df_resumen_final)} personas")
                self.logger.info(f"   📊 Total horas acumuladas: {df_resumen_final['Total Horas Trabajadas'].sum()}")
                self.logger.debug(f"   Resumen:\n{df_resumen_final.to_string()}")
                
                return df_resumen_final
            else:
                self.logger.warning("   ⚠️ No se encontraron registros completos")
                return pd.DataFrame()
            
        except Exception as e:
            self.logger.error(f"❌ Error generando resumen: {str(e)}", exc_info=True)
            return pd.DataFrame()
    
    def generar_registros_incompletos(self) -> pd.DataFrame:
        """Genera la tercera hoja: Registros incompletos"""
        try:
            self.logger.info("⚠️  Generando registros incompletos...")
            
            df = self.df_procesado.copy()
            
            incompletos = []
            
            # Agrupar por persona y fecha
            for (nombre, fecha), grupo in df.groupby(['Nombre', 'Fecha']):
                
                # Contar entradas y salidas
                entradas = grupo[grupo['Tipo Persona'] == 'entrada'].shape[0]
                salidas = grupo[grupo['Tipo Persona'] == 'salida'].shape[0]
                
                # Si falta entrada o salida
                if entradas == 0 or salidas == 0:
                    tipo_falta = []
                    
                    if entradas == 0:
                        tipo_falta.append("❌ FALTA ENTRADA")
                    if salidas == 0:
                        tipo_falta.append("❌ FALTA SALIDA")
                    
                    incompletos.append({
                        'Nombre': nombre,
                        'Fecha': fecha,
                        'Tipo de Falta': ' | '.join(tipo_falta),
                        'Sensor': str(grupo['Sensor'].values[0]).strip() if len(grupo['Sensor'].values) > 0 else 'N/A'
                    })
            
            df_incompletos = pd.DataFrame(incompletos)
            
            if len(df_incompletos) > 0:
                df_incompletos = df_incompletos.sort_values(['Nombre', 'Fecha']).reset_index(drop=True)
                self.logger.info(f"   ✅ {len(df_incompletos)} registros incompletos encontrados")
                return df_incompletos
            else:
                self.logger.info("   ✅ No hay registros incompletos")
                return pd.DataFrame()
            
        except Exception as e:
            self.logger.error(f"❌ Error generando incompletos: {str(e)}", exc_info=True)
            return pd.DataFrame()
    
    def aplicar_formato_tabla(self, archivo_excel: str):
        """Aplica formato de tabla a todas las hojas del Excel"""
        try:
            self.logger.info("🎨 Aplicando formato de tabla a las hojas...")
            
            wb = load_workbook(archivo_excel)
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Obtener rango de datos
                if ws.max_row > 1:  # Si hay datos
                    tab = Table(displayName=f"Tabla_{sheet_name.replace(' ', '_')}", 
                               ref=f"A1:{chr(64 + ws.max_column)}{ws.max_row}")
                    
                    # Aplicar estilo de tabla
                    style = TableStyleInfo(
                        name="TableStyleMedium2",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False
                    )
                    tab.tableStyleInfo = style
                    ws.add_table(tab)
                    
                    # Ajustar ancho de columnas
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    self.logger.info(f"   ✅ Formato tabla aplicado a '{sheet_name}'")
            
            wb.save(archivo_excel)
            self.logger.info(f"✅ Formato de tabla aplicado a todas las hojas")
            
        except Exception as e:
            self.logger.error(f"❌ Error aplicando formato de tabla: {str(e)}", exc_info=True)
    
    def guardar_reporte_excel(self, df_completos: pd.DataFrame, 
                             df_resumen: pd.DataFrame, 
                             df_incompletos: pd.DataFrame) -> str:
        """Guarda un ÚNICO Excel con 3 hojas en formato tabla"""
        try:
            self.logger.info("💾 Guardando reporte Excel único con formato tabla...")
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            archivo_salida = f"{self.directorio_salida}/Reporte_Biometrico_{timestamp}.xlsx"
            
            with pd.ExcelWriter(archivo_salida, engine='openpyxl') as writer:
                
                # Hoja 1: Datos Completos
                if len(df_completos) > 0:
                    df_completos.to_excel(writer, sheet_name='Datos Completos', index=False)
                    self.logger.info(f"   ✅ Hoja 1 'Datos Completos' guardada ({len(df_completos)} registros)")
                else:
                    pd.DataFrame({'Información': ['No hay datos disponibles']}).to_excel(
                        writer, sheet_name='Datos Completos', index=False
                    )
                    self.logger.info(f"   ✅ Hoja 1 'Datos Completos' guardada (vacía)")
                
                # Hoja 2: Resumen de Horas CON TOTAL ACUMULADO (SIN fila de totales)
                if len(df_resumen) > 0:
                    df_resumen.to_excel(writer, sheet_name='Resumen de Horas', index=False)
                    self.logger.info(f"   ✅ Hoja 2 'Resumen de Horas' guardada ({len(df_resumen)} personas)")
                else:
                    pd.DataFrame({'Información': ['No hay horas registradas completas']}).to_excel(
                        writer, sheet_name='Resumen de Horas', index=False
                    )
                    self.logger.info(f"   ✅ Hoja 2 'Resumen de Horas' guardada (vacía)")
                
                # Hoja 3: Registros Incompletos
                if len(df_incompletos) > 0:
                    df_incompletos.to_excel(writer, sheet_name='Registros Incompletos', index=False)
                    self.logger.info(f"   ✅ Hoja 3 'Registros Incompletos' guardada ({len(df_incompletos)} registros)")
                else:
                    pd.DataFrame({'Información': ['Todos los registros están completos']}).to_excel(
                        writer, sheet_name='Registros Incompletos', index=False
                    )
                    self.logger.info(f"   ✅ Hoja 3 'Registros Incompletos' guardada (vacía)")
            
            # Aplicar formato de tabla
            self.aplicar_formato_tabla(archivo_salida)
            
            self.archivo_excel_salida = archivo_salida
            self.logger.info(f"✅ Reporte ÚNICO guardado en: {archivo_salida}")
            print(f"\n✅ Reporte Excel ÚNICO generado: {os.path.basename(archivo_salida)}")
            
            return archivo_salida
            
        except Exception as e:
            self.logger.error(f"❌ Error al guardar: {str(e)}", exc_info=True)
            return None
    
    def ejecutar(self) -> str:
        """Ejecuta el proceso completo"""
        try:
            print("\n" + "=" * 80)
            print("🚀 GENERADOR DE REPORTE BIOMÉTRICO - EXCEL CON TABLAS")
            print("=" * 80 + "\n")
            
            # Cargar datos
            if self.cargar_datos() is None:
                self.logger.error("Abortando: Error cargando datos")
                return None
            
            # Procesar datos
            if self.procesar_datos() is None:
                self.logger.error("Abortando: Error procesando datos")
                return None
            
            # Generar las 3 hojas
            df_completos = self.generar_datos_completos()
            df_resumen = self.generar_resumen_horas()
            df_incompletos = self.generar_registros_incompletos()
            
            # Guardar único Excel con formato tabla
            archivo_reporte = self.guardar_reporte_excel(df_completos, df_resumen, df_incompletos)
            
            self.logger.info("\n" + "=" * 80)
            self.logger.info("✅ PROCESO COMPLETADO EXITOSAMENTE")
            self.logger.info("=" * 80)
            self.logger.info(f"📊 Archivo de reporte: {archivo_reporte}")
            self.logger.info(f"📝 Archivo de logs: {self.archivo_log}")
            self.logger.info("=" * 80)
            
            print("=" * 80)
            print("✅ PROCESO COMPLETADO EXITOSAMENTE")
            print("=" * 80)
            print(f"📊 Archivo Excel: {os.path.basename(archivo_reporte)}")
            print(f"📝 Archivo Logs: {os.path.basename(self.archivo_log)}")
            print("=" * 80 + "\n")
            
            return archivo_reporte
            
        except Exception as e:
            self.logger.error(f"❌ Error en el proceso: {str(e)}", exc_info=True)
            return None
    
    @staticmethod
    def _calcular_horas(hora_inicio: str, hora_fin: str) -> float:
        """Calcula diferencia de horas entre dos tiempos"""
        try:
            if not hora_inicio or not hora_fin:
                return 0
            
            dt_inicio = datetime.strptime(str(hora_inicio).strip(), '%H:%M:%S').time()
            dt_fin = datetime.strptime(str(hora_fin).strip(), '%H:%M:%S').time()
            
            fecha_temp = datetime.now().date()
            dt_inicio_full = datetime.combine(fecha_temp, dt_inicio)
            dt_fin_full = datetime.combine(fecha_temp, dt_fin)
            
            if dt_fin_full < dt_inicio_full:
                dt_fin_full += timedelta(days=1)
            
            diferencia = dt_fin_full - dt_inicio_full
            horas = diferencia.total_seconds() / 3600
            
            return max(0, horas)
        except:
            return 0


# ==========================================
# MAIN
# ==========================================

if __name__ == "__main__":
    try:
        reporte = ReporteBiometrico(directorio_trabajo=".")
        archivo = reporte.ejecutar()
        
    except KeyboardInterrupt:
        print("\n\n❌ Proceso cancelado por el usuario")
    except Exception as e:
        print(f"\n\n❌ Error: {str(e)}")