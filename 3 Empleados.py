import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from calendar import month_name
from datetime import datetime, timedelta
import re
import unicodedata
import os
import glob
import sys
from script_runtime import get_input_dir, get_output_dir, get_input_file

# Configurar salida UTF-8 para evitar errores con emojis en Windows
if sys.stdout.encoding != 'utf-8':
    try:
        sys.stdout.reconfigure(encoding='utf-8')
    except AttributeError:
        # Fallback para versiones de python < 3.7
        import codecs
        sys.stdout = codecs.getwriter("utf-8")(sys.stdout.detach())

# =========================
# CONFIGURACIÓN
# =========================
# Ruta de destino en Google Drive compartido
ruta_destino = get_output_dir(r"G:\Mi unidad\ARCHIVOS_COMPARTIDOS_LAYUNTA\SISTEMAS\INFORMES EMPLEADOS")

# Lista de sucursales (SOLO SUCURSALES para este informe)
SUCURSALES = [
    "LY01-Ballofet",
    "LY02-Velez", 
    "LY04-Alem",
    "LY07-Cuadro Benegas",
    "LY19-Alvear",
    "LY22-CENTRO",
    "LY34-Bowen",
    "LY37-Libertador",
    "LY42-Atuel Norte"
]

# =========================
# FUNCIÓN PARA ENCONTRAR ARCHIVO CSV
# =========================
def encontrar_archivo_csv():
    """
    Busca automáticamente archivos CSV en el directorio actual.
    Retorna el nombre del primer archivo CSV encontrado.
    """
    archivo_directo = get_input_file()
    if archivo_directo and archivo_directo.lower().endswith('.csv'):
        print(f"Archivo subido detectado: {os.path.basename(archivo_directo)}")
        return archivo_directo

    # Buscar todos los archivos CSV en el directorio configurado
    archivos_csv = glob.glob(os.path.join(get_input_dir(), "*.csv"))
    
    if not archivos_csv:
        print("❌ No se encontraron archivos CSV en el directorio actual")
        print("📁 Directorio actual:", os.getcwd())
        return None
    
    if len(archivos_csv) == 1:
        print(f"✅ Se encontró 1 archivo CSV: {archivos_csv[0]}")
        return archivos_csv[0]
    else:
        print("📂 Se encontraron múltiples archivos CSV. Seleccionando el más reciente automáticamente.")
        archivos_csv.sort(key=os.path.getmtime, reverse=True)
        return archivos_csv[0]

# =========================
# FUNCIÓN PARA DETERMINAR SEMANAS DEL MES
# =========================
def obtener_semana_mes(fecha, semana_inicio_dia=4):
    """
    Determina la semana del mes basado en un día de inicio específico.
    Por ejemplo, si semana_inicio_dia=4, la semana 1 empieza el día 4.
    """
    if pd.isna(fecha):
        return 0
        
    dia_mes = fecha.day
    
    if dia_mes < semana_inicio_dia:
        return 0  # Días antes de la primera semana
    
    # Calcular semana basado en el día de inicio
    semana = ((dia_mes - semana_inicio_dia) // 7) + 1
    
    # Asegurar que no exceda la semana 5
    return min(semana, 5)

def obtener_rango_semana(mes, anio, numero_semana, semana_inicio_dia=4):
    """
    Obtiene el rango de fechas para una semana específica del mes.
    """
    if numero_semana == 0:
        # Para días antes de la primera semana
        fecha_inicio = datetime(anio, mes, 1)
        fecha_fin = datetime(anio, mes, semana_inicio_dia - 1)
        return fecha_inicio, fecha_fin
    
    # Día de inicio de la semana
    dia_inicio = semana_inicio_dia + (numero_semana - 1) * 7
    fecha_inicio = datetime(anio, mes, dia_inicio)
    
    # Día de fin de la semana (6 días después)
    fecha_fin = fecha_inicio + timedelta(days=6)
    
    # Ajustar si el fin de semana excede el mes
    if mes == 12:
        ultimo_dia_mes = 31
    else:
        # Manejar cambio de año si es necesario (simplificado aquí para el mismo mes)
        try:
            proximo_mes = datetime(anio, mes + 1, 1)
        except ValueError:
            proximo_mes = datetime(anio + 1, 1, 1)
        ultimo_dia_mes = (proximo_mes - timedelta(days=1)).day
        
    if fecha_fin.day > ultimo_dia_mes:
        fecha_fin = datetime(anio, mes, ultimo_dia_mes)
    
    return fecha_inicio, fecha_fin

# =========================
# BUSCAR ARCHIVO CSV AUTOMÁTICAMENTE
# =========================
archivo_entrada = encontrar_archivo_csv()
if archivo_entrada is None:
    sys.exit(1)

# Determinar automáticamente la semana actual para procesar
hoy = datetime.today()
mes_analizar = hoy.month
anio_analizar = hoy.year
semana_analizar = obtener_semana_mes(hoy)

fecha_inicio, fecha_fin = obtener_rango_semana(mes_analizar, anio_analizar, semana_analizar)

print(f"\n📊 Analizando automáticamente semana {semana_analizar}: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}")

# =========================
# LECTURA DEL CSV
# =========================
try:
    df = pd.read_csv(archivo_entrada, encoding="utf-8", sep=';', on_bad_lines='skip')
    print(f"✅ CSV leído exitosamente. Filas: {len(df)}, Columnas: {len(df.columns)}")
except Exception as e:
    print(f"❌ Error al leer el CSV: {e}")
    try:
        df = pd.read_csv(archivo_entrada, encoding="latin-1", sep=';', on_bad_lines='skip')
        print("✅ CSV leído con encoding latin-1")
    except Exception as e2:
        print(f"❌ No se pudo leer el CSV: {e2}")
        exit()

print("Columnas originales:", df.columns.tolist())

# =========================
# NORMALIZAR NOMBRES DE COLUMNAS
# =========================
original_columns = df.columns.tolist()
new_columns = []

for i, col in enumerate(original_columns):
    if col == '$Beneficio':
        col_normalized = 'Beneficio_monetario'
    elif col == '%Beneficio':
        col_normalized = 'Beneficio_porcentaje'
    elif col == 'Beneficio':
        col_normalized = 'Beneficio_general'
    else:
        col_normalized = unicodedata.normalize('NFKD', str(col))
        col_normalized = col_normalized.encode('ascii', errors='ignore').decode('ascii')
        col_normalized = re.sub(r'[^a-zA-Z0-9]', '_', col_normalized)
        col_normalized = re.sub(r'_+', '_', col_normalized)
        col_normalized = col_normalized.strip('_')
    
    base_name = col_normalized
    counter = 1
    while col_normalized in new_columns:
        col_normalized = f"{base_name}_{counter}"
        counter += 1
    
    new_columns.append(col_normalized)

df.columns = new_columns

print("Columnas normalizadas:", df.columns.tolist())

# =========================
# IDENTIFICAR COLUMNAS CORRECTAS
# =========================
def encontrar_columna(patrones):
    for col in df.columns:
        col_lower = col.lower()
        for patron in patrones:
            if patron in col_lower:
                return col
    return None

# DIAGNÓSTICO DE COLUMNAS
print("\n🔍 DIAGNÓSTICO DE COLUMNAS:")
for col in df.columns:
    if any(term in col.lower() for term in ['nro', 'trx', 'transaccion', 'numero', 'fecha', 'beneficio', 'tienda']):
        print(f"  {col}: {df[col].dtype} - Ejemplo: {df[col].iloc[0] if len(df) > 0 else 'N/A'}")

# IDENTIFICAR COLUMNAS
col_mensaje = encontrar_columna(['mensaje', 'promocion'])

col_nro_trx = None
for col in df.columns:
    col_lower = col.lower()
    if ('nro' in col_lower or 'numero' in col_lower) and 'trx' in col_lower and 'fecha' not in col_lower:
        col_nro_trx = col
        break
if col_nro_trx is None:
    col_nro_trx = encontrar_columna(['transaccion', 'trx'])

col_fecha = encontrar_columna(['fecha', 'inicio', 'date'])
col_beneficio = encontrar_columna(['beneficio_monetario', 'dbeneficio'])
col_tienda = encontrar_columna(['tienda', 'sucursal', 'store', 'local'])

if col_beneficio is None:
    print("Buscando columna $Beneficio alternativas...")
    col_beneficio = encontrar_columna(['monetario', 'monto', 'valor', 'importe'])
    if col_beneficio is None:
        for col in df.columns:
            if 'beneficio' in col.lower() and col != 'Beneficio_porcentaje':
                col_beneficio = col
                break

# VERIFICAR COLUMNAS REQUERIDAS
columnas_requeridas = {
    'Mensaje': col_mensaje,
    'NroTrx': col_nro_trx,
    'Fecha': col_fecha,
    'Beneficio ($)': col_beneficio,
    'Tienda': col_tienda
}

columnas_faltantes = [k for k, v in columnas_requeridas.items() if v is None]
if columnas_faltantes:
    print(f"❌ No se encontraron las siguientes columnas: {columnas_faltantes}")
    print("Columnas disponibles:", df.columns.tolist())
    exit()

print(f"✅ Columna mensaje: {col_mensaje}")
print(f"✅ Columna nro trx: {col_nro_trx}")
print(f"✅ Columna fecha: {col_fecha}")
print(f"✅ Columna beneficio ($): {col_beneficio}")
print(f"✅ Columna tienda: {col_tienda}")

# =========================
# FILTRADO PROMOCIONES
# =========================
mensajes_filtrar = [
    "Empleados 15% de descuento",
]

df_filtrado = df[df[col_mensaje].str.contains('|'.join(mensajes_filtrar), case=False, na=False)].copy()

if df_filtrado.empty:
    print("⚠ No hay datos que coincidan con los mensajes de promoción")
    exit()

print(f"✅ Datos filtrados: {len(df_filtrado)} filas encontradas")

# =========================
# PROCESAMIENTO DE DATOS
# =========================
try:
    def limpiar_valor_monetario(valor):
        if pd.isna(valor):
            return 0.0
        
        valor_str = str(valor).strip()
        
        if valor_str == '' or valor_str == 'nan':
            return 0.0
        
        valor_str = valor_str.replace(',', '.')
        
        if valor_str.count('.') > 1:
            partes = valor_str.split('.')
            parte_entera = ''.join(partes[:-1])
            parte_decimal = partes[-1]
            valor_str = f"{parte_entera}.{parte_decimal}"
        
        valor_str = re.sub(r'[^\d.]', '', valor_str)
        
        if valor_str == '' or valor_str == '.':
            return 0.0
        
        try:
            return float(valor_str)
        except ValueError:
            return 0.0
    
    df_filtrado[col_beneficio] = df_filtrado[col_beneficio].apply(limpiar_valor_monetario)
    df_filtrado[col_fecha] = pd.to_datetime(df_filtrado[col_fecha], errors="coerce", dayfirst=True)
    df_filtrado = df_filtrado.dropna(subset=[col_fecha])
        
    print("✅ Datos procesados correctamente")

except Exception as e:
    print(f"❌ Error al procesar datos: {e}")
    exit()

# =========================
# FILTRAR POR SEMANA SELECCIONADA
# =========================
mes_nombre = month_name[mes_analizar].upper()

df_actual = df_filtrado[
    (df_filtrado[col_fecha] >= pd.Timestamp(fecha_inicio)) &
    (df_filtrado[col_fecha] <= pd.Timestamp(fecha_fin))
].copy()

if df_actual.empty:
    print(f"⚠ No hay datos para la semana {semana_analizar}: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}")
    exit()

print(f"✅ Datos para semana {semana_analizar}: {len(df_actual)} filas")

# =========================
# FILTRAR SOLO SUCURSALES (ELIMINADO FRANQUICIAS)
# =========================
df_sucursales = df_actual[df_actual[col_tienda].isin(SUCURSALES)].copy()

if df_sucursales.empty:
    print(f"⚠ No hay datos de SUCURSALES para la semana {semana_analizar}")
    exit()

print(f"📊 Sucursales: {len(df_sucursales)} filas, {df_sucursales[col_tienda].nunique()} tiendas")

# =========================
# FUNCIÓN PARA CREAR HOJA DE EXCEL (SOLO SUCURSALES)
# =========================
def crear_hoja_excel(ws, df_data, titulo_hoja):
    """
    Crea una hoja de Excel con el formato profesional
    """
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
    thick_border = Border(left=Side(style='medium'), right=Side(style='medium'),
                          top=Side(style='medium'), bottom=Side(style='medium'))
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    total_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    
    title_font = Font(bold=True, size=16, color="FFFFFF")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    total_font = Font(bold=True, size=11, color="000000")
    normal_font = Font(size=11)
    
    center_align = Alignment(horizontal="center", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # RESUMEN POR TIENDA
    resumen = df_data.groupby(col_tienda).agg({
        col_nro_trx: 'nunique',
        col_beneficio: "sum"
    }).reset_index()
    
    resumen["Total_Vendido"] = resumen[col_beneficio] / 0.13
    resumen.rename(columns={
        col_nro_trx: "Transacciones totales",
        col_beneficio: "Total Descuento",
        "Total_Vendido": "Venta Neta Total"
    }, inplace=True)
    
    # ENCABEZADO DEL REPORTE
    semanas_texto = ["ANTES PRIMERA", "PRIMERA", "SEGUNDA", "TERCERA", "CUARTA", "QUINTA"]
    
    ws.merge_cells('A1:D1')
    ws['A1'] = f"{titulo_hoja} - {semanas_texto[semana_analizar]} SEMANA DE {mes_nombre} {anio_analizar}"
    ws['A1'].font = title_font
    ws['A1'].alignment = center_align
    ws['A1'].fill = header_fill
    
    # Información resumen
    ws.merge_cells('A2:D2')
    total_transacciones = resumen["Transacciones totales"].sum()
    ws['A2'] = f"SUCURSALES - Total Tiendas: {len(resumen)} | Transacciones: {total_transacciones:,} | Período: {fecha_inicio.strftime('%d/%m/%Y')} - {fecha_fin.strftime('%d/%m/%Y')}"
    ws['A2'].font = Font(bold=True, size=10, color="000000")
    ws['A2'].alignment = center_align
    ws['A2'].fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    
    # Espacio
    ws.append([])
    
    # ENCABEZADOS DE COLUMNAS
    columnas_excel = ["Tienda", "Transacciones totales", "Total Descuento", "Venta Neta Total"]
    ws.append(columnas_excel)
    
    for col in range(1, 5):
        cell = ws.cell(row=4, column=col)
        cell.font = header_font
        cell.alignment = center_align
        cell.fill = header_fill
        cell.border = thin_border
    
    # DATOS POR TIENDA
    fila_inicio = 5
    for idx, row in resumen.iterrows():
        ws.cell(row=fila_inicio, column=1, value=row["Tienda"]).alignment = left_align
        ws.cell(row=fila_inicio, column=2, value=int(row["Transacciones totales"])).alignment = center_align
        ws.cell(row=fila_inicio, column=3, value=row["Total Descuento"])
        ws.cell(row=fila_inicio, column=4, value=row["Venta Neta Total"])
        fila_inicio += 1
    
    # APLICAR FORMATO
    for row in ws.iter_rows(min_row=5, max_row=fila_inicio-1, min_col=1, max_col=4):
        for cell in row:
            cell.border = thin_border
            cell.font = normal_font
            if cell.column in [3, 4]:
                cell.number_format = '"$"#,##0.00'
                cell.alignment = right_align
            elif cell.column == 2:
                cell.alignment = center_align
    
    # FILA DE TOTALES
    ws.cell(row=fila_inicio, column=1, value="TOTAL GENERAL").font = total_font
    ws.cell(row=fila_inicio, column=2, value=int(resumen["Transacciones totales"].sum())).font = total_font
    ws.cell(row=fila_inicio, column=3, value=resumen["Total Descuento"].sum()).font = total_font
    ws.cell(row=fila_inicio, column=4, value=resumen["Venta Neta Total"].sum()).font = total_font
    
    for col in range(1, 5):
        cell = ws.cell(row=fila_inicio, column=col)
        cell.fill = total_fill
        cell.border = thick_border
        cell.alignment = center_align
        if col in [3, 4]:
            cell.number_format = '"$"#,##0.00'
    
    # ESTADÍSTICAS ADICIONALES (SOLO SI HAY DATOS)
    if len(resumen) > 0:
        fila_inicio += 2
        
        ws.merge_cells(f'A{fila_inicio}:D{fila_inicio}')
        ws[f'A{fila_inicio}'] = "ESTADÍSTICAS SUCURSALES"
        ws[f'A{fila_inicio}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{fila_inicio}'].alignment = center_align
        ws[f'A{fila_inicio}'].fill = header_fill
        ws[f'A{fila_inicio}'].border = thin_border
        
        fila_inicio += 1
        ws[f'A{fila_inicio}'] = "Promedio de transacciones por tienda:"
        ws[f'B{fila_inicio}'] = resumen["Transacciones totales"].mean()
        ws[f'A{fila_inicio}'].font = Font(bold=True)
        ws[f'B{fila_inicio}'].font = Font(bold=True)
        ws[f'B{fila_inicio}'].alignment = center_align
        
        fila_inicio += 1
        ws[f'A{fila_inicio}'] = "Tienda con más transacciones:"
        ws[f'B{fila_inicio}'] = resumen.loc[resumen["Transacciones totales"].idxmax(), "Tienda"]
        ws[f'C{fila_inicio}'] = resumen["Transacciones totales"].max()
        ws[f'A{fila_inicio}'].font = Font(bold=True)
        ws[f'C{fila_inicio}'].font = Font(bold=True)
        ws[f'C{fila_inicio}'].alignment = center_align
        
        fila_inicio += 1
        ws[f'A{fila_inicio}'] = "Tienda con mayor venta neta:"
        ws[f'B{fila_inicio}'] = resumen.loc[resumen["Venta Neta Total"].idxmax(), "Tienda"]
        ws[f'C{fila_inicio}'] = resumen["Venta Neta Total"].max()
        ws[f'A{fila_inicio}'].font = Font(bold=True)
        ws[f'C{fila_inicio}'].font = Font(bold=True)
        ws[f'C{fila_inicio}'].number_format = '"$"#,##0.00'
        ws[f'C{fila_inicio}'].alignment = right_align
    
    # AJUSTES FINALES
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    
    ws.freeze_panes = 'A5'
    if len(resumen) > 0:
        ws.auto_filter.ref = f"A4:D{len(resumen) + 4}"
    
    return resumen

# =========================
# CREAR WORKBOOK (SOLO SUCURSALES)
# =========================
wb = Workbook()

# Eliminar hoja por defecto si existe
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# Crear hoja de Sucursales
ws_sucursales = wb.create_sheet("Sucursales")
resumen_sucursales = crear_hoja_excel(ws_sucursales, df_sucursales, "INFORME EMPLEADOS - SUCURSALES")

# =========================
# GUARDAR ARCHIVO EN GOOGLE DRIVE
# =========================
# Crear la carpeta de destino si no existe
if not os.path.exists(ruta_destino):
    try:
        os.makedirs(ruta_destino)
        print(f"✅ Carpeta creada: {ruta_destino}")
    except Exception as e:
        print(f"❌ No se pudo crear la carpeta de destino: {e}")
        # Guardar en directorio actual como respaldo
        ruta_destino = "."

nombre_archivo = f"Informe_Empleados_Semana_{semana_analizar}_{mes_nombre}_{anio_analizar}.xlsx"
ruta_completa = os.path.join(ruta_destino, nombre_archivo)

try:
    wb.save(ruta_completa)
    print(f"✅ Informe profesional generado: {ruta_completa}")
except Exception as e:
    print(f"❌ Error al guardar en Google Drive: {e}")
    # Guardar en directorio actual como respaldo
    wb.save(nombre_archivo)
    print(f"✅ Informe guardado en directorio actual: {nombre_archivo}")

print(f"📊 Sucursales: {len(resumen_sucursales)} tiendas, {resumen_sucursales['Transacciones totales'].sum():,} transacciones")
print(f"💰 Total Descuento: ${resumen_sucursales['Total Descuento'].sum():,.2f}")
print(f"💰 Venta Neta Total: ${resumen_sucursales['Venta Neta Total'].sum():,.2f}")
